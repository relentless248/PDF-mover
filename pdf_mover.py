import os
import re
import shutil
import sys
import json
import threading
import time
import gc
import traceback
import uuid
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from datetime import datetime
from collections import OrderedDict
from PIL import Image, ImageTk, ImageEnhance, ImageFilter
from pdf2image import convert_from_path
import pytesseract
import openpyxl
from openpyxl import Workbook

from core.settings import settings, reload_settings, CONFIG_FILE, INDEX_CACHE_FILE
from core.help_text import BUILTIN_HELP

# ========== 自动获取运行目录中的路径（不要修改）==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 依赖路径（取自 settings.deps，可配置，不再写死版本号）
TESSERACT_CMD = os.path.join(BASE_DIR, settings.deps['tesseract_dir'], 'tesseract.exe')
if not os.path.exists(TESSERACT_CMD):
    messagebox.showerror("错误", f"找不到Tesseract，请确认路径：{TESSERACT_CMD}")
    sys.exit(1)
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

TESSDATA_DIR = os.path.join(BASE_DIR, settings.deps['tesseract_dir'], 'tessdata')
if not os.path.exists(TESSDATA_DIR):
    messagebox.showerror("错误", f"找不到tessdata文件夹：{TESSDATA_DIR}\n请确保Tesseract-OCR文件夹内包含tessdata子文件夹。")
    sys.exit(1)
os.environ['TESSDATA_PREFIX'] = TESSDATA_DIR

POPPLER_PATH = os.path.join(BASE_DIR, settings.deps['poppler_dir'], 'Library', 'bin')
if not os.path.exists(POPPLER_PATH):
    messagebox.showerror("错误", f"找不到Poppler，请确认路径：{POPPLER_PATH}")
    sys.exit(1)

# ========== 预编译正则（由 settings 统一派生；配置修改后通过 reload_derived 刷新）==========
project_regex = settings.project_regex
see_regex = settings.see_regex
analyze_see_pattern = settings.analyze_see_pattern
analyze_no_see_pattern = settings.analyze_no_see_pattern
see_from_filename_regex = settings.see_from_filename_regex

def reload_derived():
    """配置变更后刷新模块级正则引用（供可视化设置保存后调用）。"""
    global project_regex, see_regex, analyze_see_pattern, analyze_no_see_pattern, see_from_filename_regex
    project_regex = settings.project_regex
    see_regex = settings.see_regex
    analyze_see_pattern = settings.analyze_see_pattern
    analyze_no_see_pattern = settings.analyze_no_see_pattern
    see_from_filename_regex = settings.see_from_filename_regex

class AutoPauseDialog(tk.Toplevel):
    """非模态暂停对话框（修复：关闭按钮回调）"""
    def __init__(self, parent, message, timeout_callback, skip_callback, timeout=20):
        super().__init__(parent)
        self.parent = parent
        self.timeout_callback = timeout_callback
        self.skip_callback = skip_callback
        self.timeout = timeout
        self.remaining = timeout

        self.title("自动运行暂停")
        self.geometry("450x200")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        # 绑定关闭窗口事件
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        tk.Label(self, text=message, wraplength=420, justify=tk.LEFT).pack(pady=10, padx=10)
        self.timer_label = tk.Label(self, text=f"剩余 {self.remaining} 秒后将自动跳过", fg="blue")
        self.timer_label.pack(pady=5)
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="立即处理", command=self.on_handle_now, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="跳过", command=self.on_skip, width=15).pack(side=tk.LEFT, padx=5)
        self.update_timer()

    def update_timer(self):
        self.remaining -= 1
        if self.remaining <= 0:
            self.destroy()
            if self.timeout_callback:
                self.timeout_callback()
        else:
            self.timer_label.config(text=f"剩余 {self.remaining} 秒后将自动跳过")
            self.after(1000, self.update_timer)

    def on_handle_now(self):
        self.destroy()

    def on_skip(self):
        self.destroy()
        if self.skip_callback:
            self.skip_callback()

    def on_close(self):
        """关闭按钮点击时，视为跳过"""
        self.destroy()
        if self.skip_callback:
            self.skip_callback()

class PDFMoverApp:
    # ========== 工具函数 ==========
    def run_with_timeout(self, func, args=(), kwargs={}, timeout=30):
        result = [None]
        exception = [None]
        is_completed = [False]

        def wrapper():
            try:
                result[0] = func(*args, **kwargs)
            except Exception as e:
                exception[0] = e
            finally:
                is_completed[0] = True

        thread = threading.Thread(target=wrapper, daemon=True)
        thread.start()
        thread.join(timeout)

        if not is_completed[0]:
            print(f"[超时警告] 操作执行超过{timeout}秒，已强制终止")
            raise TimeoutError(f"操作超时，超过{timeout}秒未完成")
        if exception[0] is not None:
            raise exception[0]
        return result[0]

    def get_windows_compatible_path(self, file_path):
        if not file_path:
            return file_path
        abs_path = os.path.abspath(file_path)
        if abs_path.startswith('\\\\'):
            return f'\\\\?\\UNC\\{abs_path[2:]}'
        else:
            return f'\\\\?\\{abs_path}'

    # ========== 原子写入函数 ==========
    def atomic_write_json(self, filepath, data):
        """原子写入 JSON 文件（临时文件+替换）"""
        temp_file = filepath + '.tmp'
        try:
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            os.replace(temp_file, filepath)
        except Exception as e:
            print(f"原子写入失败 {filepath}: {e}")
            if os.path.exists(temp_file):
                os.remove(temp_file)
            raise

    def atomic_write_excel(self, filepath, wb):
        """原子写入 Excel 文件（先保存到临时文件，再替换）"""
        temp_file = filepath + '.tmp'
        try:
            wb.save(temp_file)
            os.replace(temp_file, filepath)
        except Exception as e:
            print(f"原子写入Excel失败 {filepath}: {e}")
            if os.path.exists(temp_file):
                os.remove(temp_file)
            raise

    # ========== 路径安全校验 ==========
    def is_target_path_valid(self, target_path):
        """校验目标文件夹是否在允许的根目录范围内"""
        abs_target = os.path.abspath(target_path)
        for root_item in self.target_roots:
            root_path = root_item['path'] if isinstance(root_item, dict) else root_item
            abs_root = os.path.abspath(root_path)
            if abs_target.startswith(abs_root + os.sep) or abs_target == abs_root:
                return True
        return False

    def perform_file_operation(self, src, dst, is_copy=False):
        """增强版文件操作：区分错误类型，仅对可恢复错误重试"""
        max_retries = settings.file_ops['retries']
        for retry in range(max_retries):
            try:
                if is_copy:
                    shutil.copy2(src, dst)
                else:
                    shutil.move(src, dst)
                return
            except PermissionError as e:
                err_str = str(e).lower()
                # 仅对文件被占用的场景重试
                if ("being used by another process" in err_str or "占用" in err_str) and retry < max_retries - 1:
                    print(f"[重试] 文件被占用，第{retry+1}次重试，等待2秒... 目标: {os.path.basename(dst)}")
                    time.sleep(settings.file_ops['retry_interval_sec'])
                    continue
                else:
                    raise e
            except Exception as e:
                # 其他错误（文件不存在、路径无效、权限拒绝等）直接抛出
                raise e

    # ========== 按钮防护装饰器 ==========
    def button_guard(func):
        def wrapper(self, *args, **kwargs):
            if self.auto_timer_id:
                self.root.after_cancel(self.auto_timer_id)
                self.auto_timer_id = None
            if self.search_timer:
                self.root.after_cancel(self.search_timer)
                self.search_timer = None

            if self.processing:
                messagebox.showwarning("操作提示", "当前正在处理文件，请等待完成后再操作！")
                return
            self.processing = True

            old_auto_run = self.auto_run
            old_auto_run_paused = self.auto_run_paused
            self.auto_run = False
            self.auto_run_paused = True

            try:
                return func(self, *args, **kwargs)
            finally:
                self.auto_run = old_auto_run
                self.auto_run_paused = old_auto_run_paused
                self.processing = False
        return wrapper

    def __init__(self, root):
        self.root = root
        self.root.title("PDF自动归类工具（带人工确认）")
        self.root.geometry("1000x850")
        self.root.minsize(800, 700)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # 线程锁（保护共享变量）
        self.thread_lock = threading.Lock()

        self.status_var = tk.StringVar(value="就绪")
        self.index_progress_var = tk.StringVar(value="索引进度：0/0")
        self.index_total = 0
        self.index_scanned = 0

        self._apply_derived_regexes()

        # 缓存：key 为文件路径+修改时间
        self.ocr_cache = OrderedDict()
        self.search_cache = OrderedDict()

        self.records_folder = os.path.join(BASE_DIR, 'Records')
        os.makedirs(self.records_folder, exist_ok=True)
        self.saved_excel = None
        self.not_found_excel = None
        self.reconciliation_excel = None
        self.has_new_record = False
        self.excel_caches = {}
        self.records_since_last_save = 0

        self.target_roots = self.load_config()
        self.all_project_folders = []
        self.indexing = False
        self.index_thread = None
        self.last_index_update = time.time()
        self.index_stuck_alerted = False

        self.selected_year_var = tk.StringVar(value="全部")
        self.year_options = ["全部"] + [str(y) for y in settings.years_config.get("available", [])]

        self.load_index_cache()

        # 启动时清理临时文件夹
        self.temp_dir = os.path.join(BASE_DIR, 'Temp')
        os.makedirs(self.temp_dir, exist_ok=True)
        for temp_file in os.listdir(self.temp_dir):
            try:
                os.remove(os.path.join(self.temp_dir, temp_file))
            except:
                pass

        if not self.all_project_folders and self.target_roots:
            self.start_indexing()
        elif self.all_project_folders:
            self.status_var.set(f"已从缓存加载索引，共 {len(self.all_project_folders)} 个文件夹")

        self.folder_history = []
        self.source_folder = tk.StringVar()
        self.pdf_files = []
        self.current_pdf = None
        self.current_text = ""
        self.current_images = []
        self.matched_folders = []
        self.current_all_project_ids = []
        self.current_all_see_ids = []
        self.processing = False

        self.multi_project_var = tk.BooleanVar()
        self.no_intervene_var = tk.BooleanVar()
        self.no_ocr_var = tk.BooleanVar()

        self.auto_run = False
        self.auto_run_paused = False
        self.current_index = -1
        self.auto_timer_id = None
        self.pause_dialog = None

        self.rotation_angle = 0
        self.scale_factor = 1.0
        self.original_preview_img = None
        self.canvas_image_id = None

        self.current_search_retry = 0
        self.search_timer = None

        self.create_widgets()

        if not self.target_roots:
            self.root.after(100, self.guide_add_root)

        self.update_record_files()

    # ========== 配置与缓存 ==========
    def _apply_derived_regexes(self):
        """将 settings 派生正则同步到实例属性（配置变更后调用）。"""
        self.project_regex = settings.project_regex
        self.see_regex = settings.see_regex
        self.analyze_see_pattern = settings.analyze_see_pattern
        self.analyze_no_see_pattern = settings.analyze_no_see_pattern

    def open_settings(self):
        """打开可视化配置对话框；保存后刷新运行期参数。"""
        try:
            from core.settings_dialog import open_settings_dialog
        except ImportError:
            messagebox.showerror("错误", "未找到配置界面模块 core/settings_dialog.py")
            return
        if open_settings_dialog(self.root):
            # 配置已保存：刷新模块级与实例级派生正则
            reload_settings()
            reload_derived()
            self._apply_derived_regexes()
            self._apply_runtime_ocr()
            self.status_var.set("设置已保存并生效")

    def _apply_runtime_ocr(self):
        """OCR 语言等运行时参数若变化，更新 pytesseract 配置。"""
        if os.path.exists(TESSERACT_CMD):
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

    def show_help(self):
        """打开内置使用说明窗口。"""
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Toplevel) and "使用说明" in widget.title():
                widget.lift()
                widget.focus_force()
                return
        win = tk.Toplevel(self.root)
        win.title("使用说明")
        win.geometry("760x640")
        win.minsize(600, 480)
        win.transient(self.root)

        frame = tk.Frame(win)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        text_widget = tk.Text(frame, wrap=tk.WORD, font=("微软雅黑", 10))
        y_scroll = tk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        x_scroll = tk.Scrollbar(frame, orient=tk.HORIZONTAL, command=text_widget.xview)
        text_widget.config(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        text_widget.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        text_widget.insert("1.0", BUILTIN_HELP)
        text_widget.config(state="disabled")

        tk.Button(win, text="关闭", command=win.destroy, width=12).pack(pady=6)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # 兼容旧格式（纯路径列表）
                    roots = config.get('roots', [])
                    if not roots:
                        old_roots = config.get('target_roots', [])
                        roots = [{'path': r, 'depth': settings.cache['search_depth']} for r in old_roots]
                    return roots
            except:
                return []
        return []

    def save_config(self):
        try:
            self.atomic_write_json(CONFIG_FILE, {'roots': self.target_roots})
        except Exception as e:
            print(f"保存配置失败: {e}")

    def is_cache_expired(self):
        if not os.path.exists(INDEX_CACHE_FILE):
            return True
        cache_mtime = os.path.getmtime(INDEX_CACHE_FILE)
        if time.time() - cache_mtime > settings.cache['cache_ttl_days'] * 86400:
            return True
        try:
            with open(INDEX_CACHE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 获取缓存中的根目录信息
                cached_roots = data.get('roots', data.get('target_roots', []))
                # 提取当前根目录路径和深度信息
                current_root_keys = set(f"{r['path']}_{r.get('depth', settings.cache['search_depth'])}" for r in self.target_roots)
                # 提取缓存中的根目录路径和深度信息
                if isinstance(cached_roots[0], dict) if cached_roots else False:
                    cached_root_keys = set(f"{r['path']}_{r.get('depth', settings.cache['search_depth'])}" for r in cached_roots)
                else:
                    cached_root_keys = set(f"{r}_{settings.cache['search_depth']}" for r in cached_roots)
                if cached_root_keys != current_root_keys:
                    return True
        except:
            return True
        return False

    def load_index_cache(self):
        if self.is_cache_expired():
            if os.path.exists(INDEX_CACHE_FILE):
                os.remove(INDEX_CACHE_FILE)
            for f in os.listdir(BASE_DIR):
                if f.startswith('index_cache_chunk_'):
                    os.remove(os.path.join(BASE_DIR, f))
            with self.thread_lock:
                self.all_project_folders = []
            return

        if not os.path.exists(INDEX_CACHE_FILE):
            with self.thread_lock:
                self.all_project_folders = []
            return
        try:
            with open(INDEX_CACHE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                cached_roots = data.get('roots', data.get('target_roots', []))
                # 比较根目录配置是否相同
                current_root_keys = set(f"{r['path']}_{r.get('depth', settings.cache['search_depth'])}" for r in self.target_roots)
                if isinstance(cached_roots[0], dict) if cached_roots else False:
                    cached_root_keys = set(f"{r['path']}_{r.get('depth', settings.cache['search_depth'])}" for r in cached_roots)
                else:
                    cached_root_keys = set(f"{r}_{settings.cache['search_depth']}" for r in cached_roots)
                if cached_root_keys == current_root_keys:
                    if data.get('is_chunked', False):
                        folders = []
                        total_chunks = data.get('total_chunks', 0)
                        for i in range(total_chunks):
                            chunk_file = os.path.join(BASE_DIR, f'index_cache_chunk_{i}.json')
                            if os.path.exists(chunk_file):
                                with open(chunk_file, 'r', encoding='utf-8') as f_chunk:
                                    chunk_data = json.load(f_chunk)
                                    folders.extend(chunk_data)
                        with self.thread_lock:
                            self.all_project_folders = folders
                    else:
                        with self.thread_lock:
                            self.all_project_folders = data.get('folders', [])
        except Exception as e:
            print(f"加载索引缓存失败: {e}")
            with self.thread_lock:
                self.all_project_folders = []

    def save_index_cache(self):
        try:
            # 复制一份快照，避免在保存过程中被修改
            with self.thread_lock:
                folders_snapshot = self.all_project_folders.copy()
            if len(folders_snapshot) > settings.cache['chunk_size']:
                total_chunks = (len(folders_snapshot) + settings.cache['chunk_size'] - 1) // settings.cache['chunk_size']
                self.atomic_write_json(INDEX_CACHE_FILE, {
                    'roots': self.target_roots,
                    'total_chunks': total_chunks,
                    'is_chunked': True
                })
                for i in range(total_chunks):
                    start = i * settings.cache['chunk_size']
                    end = min((i+1)*settings.cache['chunk_size'], len(folders_snapshot))
                    chunk_data = folders_snapshot[start:end]
                    chunk_file = os.path.join(BASE_DIR, f'index_cache_chunk_{i}.json')
                    self.atomic_write_json(chunk_file, chunk_data)
            else:
                self.atomic_write_json(INDEX_CACHE_FILE, {
                    'roots': self.target_roots,
                    'folders': folders_snapshot,
                    'is_chunked': False
                })
                for f in os.listdir(BASE_DIR):
                    if f.startswith('index_cache_chunk_'):
                        os.remove(os.path.join(BASE_DIR, f))
        except Exception as e:
            print(f"保存索引缓存失败: {e}")

    def guide_add_root(self):
        # 修正提示：askdirectory 不支持多选，改为提示可以多次添加
        msg = "首次使用，请选择目标根目录（即所有项目文件夹所在的共同父目录）\n您可以通过「添加根目录」按钮多次添加多个根目录。"
        messagebox.showinfo("欢迎使用", msg)
        self.add_target_root()

    # ========== 动态记录文件管理 ==========
    def update_record_files(self):
        for filepath, (wb, _) in list(self.excel_caches.items()):
            try:
                wb.close()
                print(f"[缓存清理] 已关闭旧的工作簿：{filepath}")
            except:
                pass
        self.excel_caches.clear()

        src = self.source_folder.get().strip()
        if not src:
            safe_name = "未选择源文件夹"
        else:
            folder_name = os.path.basename(src.rstrip('/\\'))
            if not folder_name:
                folder_name = "根目录"
            safe_name = re.sub(r'[\\/*?:"<>|]', '_', folder_name)
        self.saved_excel = os.path.join(self.records_folder, f'已找到_{safe_name}.xlsx')
        self.not_found_excel = os.path.join(self.records_folder, f'未找到_{safe_name}.xlsx')
        self.reconciliation_excel = os.path.join(self.records_folder, f'多项目对账_{safe_name}.xlsx')
        self.ensure_excel_files()

    # ========== Excel记录 ==========
    def ensure_excel_files(self):
        if self.saved_excel is None or self.not_found_excel is None:
            self.update_record_files()
        saved_headers = ['操作时间', '源文件', '项目编号', '目标文件夹', '操作类型']
        if not os.path.exists(self.saved_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = '已录入'
            ws.append(saved_headers)
            self.atomic_write_excel(self.saved_excel, wb)
        not_found_headers = ['操作时间', '原文件名', '重命名后文件名', '项目编号']
        if not os.path.exists(self.not_found_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = '未找到'
            ws.append(not_found_headers)
            self.atomic_write_excel(self.not_found_excel, wb)

        wb1 = openpyxl.load_workbook(self.saved_excel)
        ws1 = wb1.active
        self.excel_caches[self.saved_excel] = (wb1, ws1)

        wb2 = openpyxl.load_workbook(self.not_found_excel)
        ws2 = wb2.active
        self.excel_caches[self.not_found_excel] = (wb2, ws2)

        # 对账表
        if self.reconciliation_excel is None:
            return
        reconciliation_headers = ['操作时间', '源文件名', '本次识别到的所有项目号', '当前处理的项目号', '处理状态', '备注/失败原因']
        if not os.path.exists(self.reconciliation_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = '多项目对账'
            ws.append(reconciliation_headers)
            self.atomic_write_excel(self.reconciliation_excel, wb)

        if self.reconciliation_excel not in self.excel_caches:
            wb_recon = openpyxl.load_workbook(self.reconciliation_excel)
            ws_recon = wb_recon.active
            self.excel_caches[self.reconciliation_excel] = (wb_recon, ws_recon)

    def append_record(self, filepath, data):
        try:
            wb, ws = self.excel_caches.get(filepath)
            if wb is None:
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                self.excel_caches[filepath] = (wb, ws)
            ws.append(data)
            self.has_new_record = True
            self.records_since_last_save += 1
            if self.records_since_last_save >= settings.records['auto_save_count']:
                self.save_excel_files()
                self.records_since_last_save = 0
        except Exception as e:
            print(f"记录失败: {e}")

    def save_excel_files(self):
        for filepath, (wb, _) in list(self.excel_caches.items()):
            try:
                self.atomic_write_excel(filepath, wb)
                print(f"已自动保存: {filepath}")
            except PermissionError:
                response = messagebox.askyesno("文件被占用",
                    f"文件 {os.path.basename(filepath)} 正在被 Excel 打开，无法保存。\n"
                    "是否生成一个备份文件（文件名将添加“_备份”后缀）？\n"
                    "选择“是”生成备份，选择“否”放弃本次保存。")
                if response:
                    base, ext = os.path.splitext(filepath)
                    backup_path = f"{base}_备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                    try:
                        self.atomic_write_excel(backup_path, wb)
                        messagebox.showinfo("备份已保存", f"数据已保存到备份文件：\n{backup_path}")
                    except Exception as e2:
                        messagebox.showerror("保存失败", f"无法保存备份文件：{e2}")
            except Exception as e:
                print(f"保存 {filepath} 失败: {e}")

    # ========== 年份提取与处理 ==========
    def extract_year_from_folder_name(self, folder_name):
        match = self.project_regex.search(folder_name)
        if not match:
            return None
        project_id = match.group(0)
        return self.extract_year_from_project_id(project_id)

    def extract_year_from_project_id(self, project_id):
        if not project_id:
            return None
        prefix_len = settings.year_rules.get('prefix_length', 4)
        century = settings.year_rules.get('century_prefix', '20')
        number_part = project_id[prefix_len:] if len(project_id) > prefix_len else project_id
        if len(number_part) >= 4 and number_part[:2] == century:
            return number_part[2:4]
        elif len(number_part) >= 2:
            return number_part[:2]
        return None

    def on_year_changed(self, event):
        self.status_var.set(f"已切换年份为：{self.selected_year_var.get()}，请点击「重建索引」以更新文件夹列表")

    def extract_see_from_filename(self, filename):
        match = settings.see_from_filename_regex.search(filename)
        return match.group(1).strip() if match else None

    def extract_all_project_ids(self, text):
        matches = self.project_regex.findall(text)
        return list(OrderedDict.fromkeys([m.strip() for m in matches if m.strip()]))

    def extract_all_see_ids(self, text):
        matches = self.see_regex.findall(text)
        return list(OrderedDict.fromkeys([m.strip() for m in matches if m.strip()]))

    def analyze_remaining_files(self):
        year_stats = {}
        see_related_count = 0
        no_see_count = 0
        no_id_count = 0
        for pdf_path in self.pdf_files:
            filename = os.path.basename(pdf_path)
            project_id = None
            see_match = self.analyze_see_pattern.match(filename)
            if see_match:
                project_id = see_match.group(2)
                see_related_count += 1
            else:
                no_see_match = self.analyze_no_see_pattern.match(filename)
                if no_see_match:
                    project_id = no_see_match.group(1)
                    no_see_count += 1
                else:
                    fallback_match = self.project_regex.search(filename)
                    if fallback_match:
                        project_id = fallback_match.group(0)
                        if '关联' in filename and '未找到项目' in filename:
                            see_related_count += 1
                        elif '未找到项目' in filename:
                            no_see_count += 1
                        else:
                            no_id_count += 1
                    else:
                        no_id_count += 1
            if project_id:
                year = self.extract_year_from_project_id(project_id)
                if year:
                    year_stats[year] = year_stats.get(year, 0) + 1
        return year_stats, see_related_count, no_see_count, no_id_count

    # ========== 索引构建（线程安全 + UI 更新 via after）==========
    def start_indexing(self):
        if self.indexing and self.index_thread and self.index_thread.is_alive():
            self.indexing = False
            self.index_thread.join(timeout=3)

        with self.thread_lock:
            self.search_cache.clear()
            self.ocr_cache.clear()
            self.all_project_folders = []

        self.indexing = True
        self.last_index_update = time.time()
        self.index_stuck_alerted = False
        self.status_var.set("正在构建文件夹索引，请稍候... 已扫描: 0")
        self.index_thread = threading.Thread(target=self.build_folder_index_thread, daemon=True)
        self.index_thread.start()
        self.monitor_indexing()

    def monitor_indexing(self):
        if not self.indexing:
            return
        now = time.time()
        if not self.index_stuck_alerted and now - self.last_index_update > 30:
            self.index_stuck_alerted = True
            self.root.after(0, self.prompt_index_stuck)
        else:
            self.root.after(5000, self.monitor_indexing)

    def prompt_index_stuck(self):
        if not self.indexing:
            return
        messagebox.showinfo("索引可能卡住",
            "索引构建长时间没有进度，可能是遇到了无法访问的文件夹。\n"
            "后台仍在继续尝试，您也可以点击“重建索引”按钮重新开始。")
        self.root.after(5000, self.monitor_indexing)

    def update_index_progress_display(self, scanned, total):
        if total == 0:
            self.index_progress_var.set(f"索引进度：0/0")
        else:
            self.index_progress_var.set(f"索引进度：{scanned}/{total}")
        self.status_var.set(f"正在构建索引... 已找到 {scanned} 个符合条件的项目文件夹（共 {total} 个）")

    def scan_folder_recursive(self, current_path, current_depth, max_depth, target_year, skip_dirs):
        """递归扫描文件夹，支持指定深度"""
        folders = []
        if current_depth > max_depth:
            return folders
        try:
            with os.scandir(current_path) as entries:
                for entry in entries:
                    try:
                        if not entry.is_dir() or entry.name.startswith('.') or entry.name in skip_dirs:
                            continue
                        full_path = entry.path
                        # 检查是否是项目文件夹
                        if self.project_regex.search(entry.name):
                            folder_year = self.extract_year_from_folder_name(entry.name)
                            if folder_year:
                                self.encountered_years.add(folder_year)
                            if target_year == "全部" or (folder_year and folder_year == target_year):
                                folders.append(full_path)
                            else:
                                self.filtered_count += 1
                        # 递归扫描下一层（如果还没达到最大深度）
                        if current_depth < max_depth:
                            sub_folders = self.scan_folder_recursive(full_path, current_depth + 1, max_depth, target_year, skip_dirs)
                            folders.extend(sub_folders)
                    except Exception:
                        self.error_count += 1
                        continue
        except Exception as e:
            self.error_count += 1
            print(f"遍历目录 {current_path} 时出错: {e}")
        return folders

    def build_folder_index_thread(self):
        # 扫描过程不操作UI，只收集数据，最后一次性通过after更新
        local_folders = []
        count = 0
        self.filtered_count = 0
        self.error_count = 0
        UPDATE_INTERVAL = 100
        skip_dirs = set(settings.file_ops['skip_dirs'])

        target_year = self.selected_year_var.get()
        self.encountered_years = set()

        # 第一步：统计总数（使用递归扫描，每个根目录独立深度）
        total = 0
        for root_item in self.target_roots:
            root_dir = root_item['path']
            scan_depth = root_item.get('depth', settings.cache['search_depth'])
            if root_dir.startswith('\\\\') or not os.path.exists(root_dir):
                continue
            sub_folders = self.scan_folder_recursive(root_dir, 1, scan_depth, target_year, skip_dirs)
            total += len(sub_folders)

        self.root.after(0, self.update_index_progress_display, 0, total)

        # 第二步：扫描并收集（使用递归扫描，每个根目录独立深度）
        for root_item in self.target_roots:
            root_dir = root_item['path']
            scan_depth = root_item.get('depth', settings.cache['search_depth'])
            if root_dir.startswith('\\\\'):
                self.root.after(0, lambda: self.status_var.set(f"跳过网络路径：{root_dir}，建议复制到本地处理"))
                continue
            if not os.path.exists(root_dir):
                continue
            sub_folders = self.scan_folder_recursive(root_dir, 1, scan_depth, target_year, skip_dirs)
            for folder in sub_folders:
                local_folders.append(folder)
                count += 1
                if count % UPDATE_INTERVAL == 0:
                    self.root.after(0, self.update_index_progress_display, count, total)

        # 更新共享变量（加锁）
        with self.thread_lock:
            self.all_project_folders = local_folders
            # 更新年份选项
            if self.encountered_years:
                years_list = sorted(list(self.encountered_years))
                self.root.after(0, lambda: self._update_year_combo(years_list))

        self.root.after(0, self.update_index_progress_display, total, total)
        self.root.after(0, self.on_indexing_finished_with_filter, count, self.filtered_count, self.error_count, self.encountered_years)

    def _update_year_combo(self, years_list):
        current_values = list(self.year_combo['values'])
        new_values = ["全部"] + sorted(list(set(current_values + years_list) - {"全部"}))
        self.year_combo['values'] = new_values

    def on_indexing_finished_with_filter(self, count, filtered_count, error_count, encountered_years):
        self.indexing = False
        self.save_index_cache()

        target_year = self.selected_year_var.get()

        msg_lines = []
        msg_lines.append(f"索引构建完成！")
        msg_lines.append(f"本次共索引：{count} 个文件夹")

        if target_year != "全部":
            msg_lines.append(f"当前分拣年份：{target_year}")
            msg_lines.append(f"过滤其他年份：{filtered_count} 个")

        if error_count > 0:
            msg_lines.append(f"忽略错误：{error_count} 个")

        if encountered_years:
            sorted_years = sorted(list(encountered_years))
            years_str = "、".join(sorted_years)
            msg_lines.append("")
            msg_lines.append(f"📅 扫描发现以下年份的文件夹：")
            msg_lines.append(f"   {years_str}")

        final_msg = "\n".join(msg_lines)
        self.status_var.set(f"索引构建完成，共 {count} 个文件夹")
        self.index_progress_var.set(f"索引进度：{self.index_total}/{self.index_total}（完成）")
        messagebox.showinfo("索引完成", final_msg)
        gc.collect()

    # ========== 目标范围管理 ==========
    @button_guard
    def add_target_root(self, allow_multiple=False):
        folder = filedialog.askdirectory(title="选择目标根目录")
        if folder:
            existing_paths = [r['path'] for r in self.target_roots]
            if folder not in existing_paths:
                # 弹出扫描深度设置对话框
                depth_dialog = tk.Toplevel(self.root)
                depth_dialog.title("设置扫描深度")
                depth_dialog.geometry("360x180")
                depth_dialog.transient(self.root)
                depth_dialog.grab_set()
                
                tk.Label(depth_dialog, text="请选择该根目录的扫描深度：").pack(pady=10)
                depth_var = tk.StringVar(value="1")
                depth_combo = ttk.Combobox(depth_dialog, textvariable=depth_var, 
                                           values=["1", "2", "3", "4", "5"], 
                                           state="normal", width=10)
                depth_combo.pack(pady=5)
                depth_combo.current(0)
                
                def confirm_depth():
                    try:
                        depth = int(depth_var.get())
                    except:
                        depth = settings.cache['search_depth']
                    self.target_roots.append({'path': folder, 'depth': depth})
                    display_text = f"{folder:.<60}  [{depth}层]"
                    self.roots_listbox.insert(tk.END, display_text)
                    self.save_config()
                    depth_dialog.destroy()
                    self.start_indexing()
                
                btn_frame = tk.Frame(depth_dialog)
                btn_frame.pack(pady=10)
                tk.Button(btn_frame, text="确定", command=confirm_depth, width=10).pack(side=tk.LEFT, padx=5)
                tk.Button(btn_frame, text="取消", command=depth_dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)
            else:
                messagebox.showinfo("提示", "该根目录已在列表中")

    @button_guard
    def edit_root_depth(self):
        selection = self.roots_listbox.curselection()
        if not selection:
            messagebox.showwarning("提示", "请先选中要编辑的根目录")
            return
        index = selection[0]
        root_item = self.target_roots[index]
        root_path = root_item['path']
        current_depth = root_item.get('depth', settings.cache['search_depth'])
        
        # 弹出扫描深度编辑对话框
        depth_dialog = tk.Toplevel(self.root)
        depth_dialog.title("编辑扫描深度")
        depth_dialog.geometry("360x180")
        depth_dialog.transient(self.root)
        depth_dialog.grab_set()
        
        tk.Label(depth_dialog, text=f"根目录：{root_path}").pack(pady=5)
        tk.Label(depth_dialog, text="请选择新的扫描深度：").pack(pady=5)
        depth_var = tk.StringVar(value=str(current_depth))
        depth_combo = ttk.Combobox(depth_dialog, textvariable=depth_var, 
                                   values=["1", "2", "3", "4", "5"], 
                                   state="normal", width=10)
        depth_combo.pack(pady=5)
        depth_combo.current(current_depth - 1)
        
        def confirm_edit():
            try:
                new_depth = int(depth_var.get())
            except:
                new_depth = settings.cache['search_depth']
            self.target_roots[index]['depth'] = new_depth
            # 更新列表显示
            self.roots_listbox.delete(index)
            display_text = f"{root_path:.<60}  [{new_depth}层]"
            self.roots_listbox.insert(index, display_text)
            self.save_config()
            depth_dialog.destroy()
            # 清理缓存，下次重建索引时使用新深度
            try:
                if os.path.exists(INDEX_CACHE_FILE):
                    os.remove(INDEX_CACHE_FILE)
                for f in os.listdir(BASE_DIR):
                    if f.startswith('index_cache_chunk_'):
                        os.remove(os.path.join(BASE_DIR, f))
            except:
                pass
        
        btn_frame = tk.Frame(depth_dialog)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="确定", command=confirm_edit, width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="取消", command=depth_dialog.destroy, width=10).pack(side=tk.LEFT, padx=5)

    @button_guard
    def remove_target_root(self):
        selection = self.roots_listbox.curselection()
        if not selection:
            messagebox.showwarning("提示", "请先选中要删除的根目录")
            return
        index = selection[0]
        root_item_to_delete = self.target_roots[index]
        root_path = root_item_to_delete['path']
        confirm = messagebox.askyesno("确认删除", 
            f"确定要删除根目录：\n{root_path}\n\n删除后将自动清理索引缓存并重建。")
        if not confirm:
            return
        self.target_roots.pop(index)
        self.roots_listbox.delete(index)
        self.save_config()
        try:
            if os.path.exists(INDEX_CACHE_FILE):
                os.remove(INDEX_CACHE_FILE)
            for f in os.listdir(BASE_DIR):
                if f.startswith('index_cache_chunk_'):
                    os.remove(os.path.join(BASE_DIR, f))
        except Exception as e:
            print(f"[缓存清理] 删除缓存文件失败：{e}")
        with self.thread_lock:
            self.all_project_folders = []
            self.search_cache.clear()
            self.ocr_cache.clear()
        self.start_indexing()
        messagebox.showinfo("删除成功", f"根目录已删除，正在重建索引...")

    @button_guard
    def refresh_index(self):
        try:
            if os.path.exists(INDEX_CACHE_FILE):
                os.remove(INDEX_CACHE_FILE)
            for f in os.listdir(BASE_DIR):
                if f.startswith('index_cache_chunk_'):
                    os.remove(os.path.join(BASE_DIR, f))
        except Exception as e:
            print(f"[缓存清理] 删除缓存失败：{e}")
        with self.thread_lock:
            self.search_cache.clear()
            self.ocr_cache.clear()
        self.start_indexing()

    @button_guard
    def refresh_selected_root(self):
        selection = self.roots_listbox.curselection()
        if not selection:
            messagebox.showwarning("提示", "请先选中要刷新的根目录")
            return
        index = selection[0]
        root_item = self.target_roots[index]
        root_dir = root_item['path']
        scan_depth = root_item.get('depth', settings.cache['search_depth'])
        if not os.path.exists(root_dir):
            messagebox.showerror("错误", f"根目录不存在：{root_dir}")
            return
        self.status_var.set(f"正在刷新根目录: {root_dir} (深度:{scan_depth}层)")
        self.root.update()
        new_folders = []
        try:
            skip_dirs = set(settings.file_ops['skip_dirs'])
            new_folders = self.scan_folder_recursive(root_dir, 1, scan_depth, "全部", skip_dirs)
        except Exception as e:
            messagebox.showerror("刷新失败", str(e))
            return
        with self.thread_lock:
            self.all_project_folders = [f for f in self.all_project_folders if not f.startswith(root_dir)]
            self.all_project_folders.extend(new_folders)
        self.save_index_cache()
        self.status_var.set(f"根目录刷新完成，新增 {len(new_folders)} 个文件夹")
        messagebox.showinfo("刷新完成", f"根目录 {root_dir} (深度:{scan_depth}层) 已刷新，当前共 {len(self.all_project_folders)} 个文件夹")

    # ========== 手动选择目标文件夹 ==========
    def browse_target_folder(self):
        folder = filedialog.askdirectory(title="选择目标文件夹（可以是项目文件夹或Test Data子文件夹）")
        if folder:
            # 安全校验
            if not self.is_target_path_valid(folder):
                messagebox.showerror("错误", "选择的文件夹不在允许的根目录范围内！")
                return
            self.folder_combo.set(folder)
            if folder not in self.folder_history:
                self.folder_history.append(folder)
            self.folder_combo['values'] = self.folder_history
            self.status_var.set(f"已手动选择: {folder}")

    # ========== 缩放与拖动 ==========
    def zoom_in(self):
        if self.original_preview_img:
            self.scale_factor = min(self.scale_factor * 1.2, 3.0)
            self.update_canvas_image()

    def zoom_out(self):
        if self.original_preview_img:
            self.scale_factor = max(self.scale_factor / 1.2, 0.5)
            self.update_canvas_image()

    def update_canvas_image(self):
        if not self.original_preview_img:
            return
        orig_w, orig_h = self.original_preview_img.size
        new_w = int(orig_w * self.scale_factor)
        new_h = int(orig_h * self.scale_factor)
        resized = self.original_preview_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        self.tk_image = ImageTk.PhotoImage(resized)
        self.canvas.delete("all")
        self.canvas_image_id = self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_image)
        self.canvas.config(scrollregion=self.canvas.bbox("all"))

    def on_drag_start(self, event):
        self.canvas.scan_mark(event.x, event.y)

    def on_drag_move(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def on_ctrl_mousewheel(self, event):
        if hasattr(event, 'delta'):
            if event.delta > 0:
                self.zoom_in()
            else:
                self.zoom_out()
        elif hasattr(event, 'num'):
            if event.num == 4:
                self.zoom_in()
            elif event.num == 5:
                self.zoom_out()

    # ========== 自动运行功能 ==========
    def cancel_auto_timer(self):
        if self.auto_timer_id:
            self.root.after_cancel(self.auto_timer_id)
            self.auto_timer_id = None

    def close_pause_dialog(self):
        if self.pause_dialog and self.pause_dialog.winfo_exists():
            self.pause_dialog.destroy()
        self.pause_dialog = None

    def start_auto_run(self):
        if self.auto_run:
            return
        if not self.pdf_files:
            messagebox.showwarning("提示", "当前源文件夹中没有PDF文件，请先选择源文件夹")
            return
        if not self.all_project_folders:
            messagebox.showwarning("提示", "项目文件夹索引为空，请先添加目标根目录并等待索引构建完成")
            return
        self.auto_run = True
        self.auto_run_paused = False
        self.processing = False
        self.current_index = 0
        self.status_var.set("自动运行模式已启动，将按顺序处理文件...")
        self.btn_auto_run.config(text="停止自动", command=self.stop_auto_run, bg='lightcoral')
        self.root.after(200, self.auto_run_next)

    def stop_auto_run(self):
        self.cancel_auto_timer()
        self.close_pause_dialog()
        self.auto_run = False
        self.auto_run_paused = False
        self.status_var.set("自动运行已停止")
        self.btn_auto_run.config(text="自动运行", command=self.start_auto_run, bg='lightgreen')

    def continue_auto_run(self):
        self.cancel_auto_timer()
        self.close_pause_dialog()
        if self.auto_run_paused:
            self.auto_run_paused = False
            self.status_var.set("自动运行已恢复")
            self.root.after(100, self.auto_run_next)

    def auto_run_next(self):
        if self.processing:
            print(f"[自动运行] 文件处理中，1秒后重试...")
            self.root.after(1000, self.auto_run_next)
            return
        if not self.auto_run or self.auto_run_paused:
            return
        if self.current_index >= len(self.pdf_files):
            self.stop_auto_run()
            year_stats, see_related_count, no_see_count, no_id_count = self.analyze_remaining_files()
            msg = "✅ 自动运行处理完毕！\n\n"
            msg += f"剩余未处理文件总数：{len(self.pdf_files)}\n"
            msg += f"   - 有关联号文件：{see_related_count}\n"
            msg += f"   - 无关联号文件：{no_see_count}\n"
            msg += f"   - 未识别编号文件：{no_id_count}\n"
            if year_stats:
                msg += "\n📅 剩余文件按年份统计：\n"
                for year in sorted(year_stats.keys()):
                    msg += f"   {year}年：{year_stats[year]} 个文件\n"
            messagebox.showinfo("处理完成统计", msg)
            return
        target_file = self.pdf_files[self.current_index]
        self.listbox.selection_clear(0, tk.END)
        self.listbox.selection_set(self.current_index)
        self.listbox.activate(self.current_index)
        self.on_select_pdf(None)

    def show_pause_dialog(self, message, skip_callback):
        self.auto_run_paused = True
        self.cancel_auto_timer()
        self.close_pause_dialog()
        self.pause_dialog = AutoPauseDialog(self.root, message, skip_callback, skip_callback, 20)
        self.auto_timer_id = self.root.after(20000, skip_callback)

    # ========== 安全PDF路径处理（临时文件+自动清理）==========
    def get_safe_pdf_path(self, pdf_path):
        needs_temp = False
        try:
            pdf_path.encode('ascii')
        except UnicodeEncodeError:
            needs_temp = True
        if pdf_path.startswith('\\\\') or pdf_path.startswith('//'):
            needs_temp = True
        if not needs_temp:
            return pdf_path, False
        safe_filename = f"temp_{uuid.uuid4().hex}.pdf"
        safe_path = os.path.join(self.temp_dir, safe_filename)
        try:
            shutil.copy2(pdf_path, safe_path)
            return safe_path, True
        except Exception as e:
            print(f"[警告] 无法创建临时文件，尝试直接处理: {e}")
            return pdf_path, False

    # ========== 自动复制到多个文件夹 ==========
    def auto_copy_to_multiple_folders(self, pdf_path, project_id, matched_folders):
        try:
            copied_count = 0
            error_folders = []
            target_testdata_folders = []
            for folder in matched_folders:
                testdata_folders = self.find_test_data_folders(folder)
                if testdata_folders:
                    target_testdata_folders.extend(testdata_folders)
            if not target_testdata_folders:
                self.auto_skip_no_folder_keep(pdf_path, project_id)
                return
            for testdata_folder in target_testdata_folders:
                # 安全校验
                if not self.is_target_path_valid(testdata_folder):
                    error_folders.append((testdata_folder, "目标文件夹不在允许范围内"))
                    continue
                target_path = os.path.join(testdata_folder, f"{settings.target_filename}.pdf")
                counter = 1
                while os.path.exists(target_path):
                    target_path = os.path.join(testdata_folder, f"{settings.target_filename}_{counter}.pdf")
                    counter += 1
                try:
                    self.perform_file_operation(pdf_path, target_path, is_copy=True)
                    copied_count += 1
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.append_record(self.saved_excel, [timestamp, os.path.basename(pdf_path),
                                                           project_id, testdata_folder, '复制（多目标）'])
                except Exception as e:
                    error_folders.append((testdata_folder, str(e)))
            if copied_count == 0:
                raise Exception(f"所有目标 Test Data 文件夹复制均失败: {error_folders}")
            if error_folders:
                warn_msg = f"部分复制成功，但有 {len(error_folders)} 个目标复制失败：\n"
                for folder, err in error_folders:
                    warn_msg += f"{folder}: {err}\n"
                if not self.no_intervene_var.get():
                    messagebox.showwarning("复制部分失败", warn_msg)
            if not error_folders:
                try:
                    os.remove(pdf_path)
                except Exception as e:
                    if not self.no_intervene_var.get():
                        messagebox.showerror("删除失败", f"无法删除原文件 {os.path.basename(pdf_path)}: {e}")
            if pdf_path == self.current_pdf:
                self.pdf_files.pop(self.current_index)
                self.listbox.delete(self.current_index)
                self.current_pdf = None
                self.canvas.delete("all")
                self.original_preview_img = None
                self.folder_combo.set('')
                self.correct_entry.delete(0, tk.END)
                self.status_var.set(f"已自动复制到 {copied_count}/{len(target_testdata_folders)} 个 Test Data 文件夹")
            else:
                try:
                    idx = self.pdf_files.index(pdf_path)
                    self.pdf_files.pop(idx)
                    self.listbox.delete(idx)
                except ValueError:
                    pass
            self.handle_auto_run_after_action(removed=True)
        except Exception as e:
            print(f"[错误] 自动复制到多文件夹失败: {traceback.format_exc()}")
            if self.no_intervene_var.get():
                self.auto_skip_no_folder_keep(pdf_path, project_id)
            else:
                self.auto_run_paused = True
                messagebox.showerror("自动复制失败", f"处理文件 {os.path.basename(pdf_path)} 时出错：{str(e)}\n自动运行已暂停。")

    # ========== 重命名函数 ==========
    def auto_rename_and_record(self, pdf_path, new_name=None, project_id=None, reason='no_id', see_id=None, remove=True):
        try:
            old_path = pdf_path
            old_name = os.path.basename(old_path)
            dir_name = os.path.dirname(old_path)
            if new_name:
                final_name = new_name
            elif reason == 'no_id':
                final_name = settings.build_filename('unidentified', original=old_name)
            elif reason == 'no_folder' and project_id:
                final_name = settings.build_filename('not_found', project_id=project_id)
            elif reason == 'no_testdata_with_see' and project_id and see_id:
                final_name = settings.build_filename('see_relation', see_id=see_id, project_id=project_id)
            else:
                final_name = settings.build_filename('unidentified', original=old_name)
            new_path = os.path.join(dir_name, final_name)
            # 安全校验：重命名后的文件仍在源文件夹内，无需额外校验
            counter = 1
            base, ext = os.path.splitext(new_path)
            while os.path.exists(new_path):
                new_path = os.path.join(dir_name, f"{base}_{counter}{ext}")
                counter += 1
            max_retries = settings.file_ops['retries']
            rename_success = False
            last_error = None
            for retry in range(max_retries):
                try:
                    os.rename(old_path, new_path)
                    rename_success = True
                    break
                except PermissionError as e:
                    last_error = e
                    if "being used by another process" in str(e).lower() and retry < max_retries - 1:
                        time.sleep(settings.file_ops['retry_interval_sec'])
                        continue
                    else:
                        raise e
            if not rename_success:
                if self.no_intervene_var.get():
                    print(f"[不干预模式] 重命名失败，跳过文件: {old_name}")
                    if remove and pdf_path == self.current_pdf:
                        self.pdf_files.pop(self.current_index)
                        self.listbox.delete(self.current_index)
                        self.current_pdf = None
                        self.canvas.delete("all")
                        self.original_preview_img = None
                        self.folder_combo.set('')
                        self.correct_entry.delete(0, tk.END)
                    if self.auto_run:
                        self.handle_auto_run_after_action(removed=remove)
                    return
                else:
                    raise last_error
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.append_record(self.not_found_excel, [timestamp, old_name, final_name, project_id or ''])
            if remove:
                if pdf_path == self.current_pdf:
                    self.pdf_files.pop(self.current_index)
                    self.listbox.delete(self.current_index)
                    self.current_pdf = None
                    self.canvas.delete("all")
                    self.original_preview_img = None
                    self.folder_combo.set('')
                    self.correct_entry.delete(0, tk.END)
                    self.status_var.set(f"已自动重命名: {final_name}")
                else:
                    try:
                        idx = self.pdf_files.index(old_path)
                        self.pdf_files.pop(idx)
                        self.listbox.delete(idx)
                    except:
                        pass
            else:
                if pdf_path == self.current_pdf:
                    idx = self.current_index
                    self.pdf_files[idx] = new_path
                    self.listbox.delete(idx)
                    self.listbox.insert(idx, final_name)
                    self.current_pdf = new_path
                    self.status_var.set(f"已自动重命名: {final_name}")
                else:
                    try:
                        idx = self.pdf_files.index(old_path)
                        self.pdf_files[idx] = new_path
                        self.listbox.delete(idx)
                        self.listbox.insert(idx, final_name)
                    except:
                        pass
            if self.auto_run:
                self.handle_auto_run_after_action(removed=remove)
        except Exception as e:
            if not self.no_intervene_var.get():
                messagebox.showerror("自动重命名失败", str(e))
            self.auto_run_paused = not self.no_intervene_var.get()

    def auto_skip_no_id(self, pdf_path):
        self.close_pause_dialog()
        if not self.auto_run or self.auto_run_paused:
            self.auto_run_paused = False
        self.auto_rename_and_record(pdf_path, reason='no_id', remove=True)

    def auto_skip_no_folder(self, pdf_path, project_id):
        self.close_pause_dialog()
        if not self.auto_run or self.auto_run_paused:
            self.auto_run_paused = False
        self.auto_rename_and_record(pdf_path, project_id=project_id, reason='no_folder', remove=True)

    def auto_skip_no_folder_keep(self, pdf_path, project_id):
        self.close_pause_dialog()
        if not self.auto_run or self.auto_run_paused:
            self.auto_run_paused = False
        self.auto_rename_and_record(pdf_path, project_id=project_id, reason='no_folder', remove=False)

    def auto_skip_no_testdata_with_see(self, pdf_path, project_id, see_id):
        self.close_pause_dialog()
        if not self.auto_run or self.auto_run_paused:
            self.auto_run_paused = False
        self.auto_rename_and_record(pdf_path, project_id=project_id, see_id=see_id, reason='no_testdata_with_see', remove=False)

    # ========== 处理 see 关联归档 ==========
    def auto_handle_see_relation(self, pdf_path, original_id, see_id):
        try:
            print(f"[see关联] 处理原项目 {original_id} 的关联项目 {see_id}")
            matches = []
            if see_id in self.search_cache:
                matches = self.search_cache[see_id]
            else:
                with self.thread_lock:
                    folders = self.all_project_folders.copy()
                for folder in folders:
                    if self.folder_matches(os.path.basename(folder), see_id):
                        matches.append(folder)
                self.search_cache[see_id] = matches
                self.search_cache.move_to_end(see_id)
                if len(self.search_cache) > settings.cache['search_max']:
                    self.search_cache.popitem(last=False)
            if len(matches) != 1:
                print(f"[see关联] 关联项目 {see_id} 匹配到 {len(matches)} 个文件夹，归档失败")
                self.auto_skip_no_testdata_with_see(pdf_path, original_id, see_id)
                return
            project_folder = matches[0]
            test_data_folders = self.find_test_data_folders(project_folder)
            if len(test_data_folders) != 1:
                print(f"[see关联] 关联项目 {see_id} 下找到 {len(test_data_folders)} 个Test Data，归档失败")
                self.auto_skip_no_testdata_with_see(pdf_path, original_id, see_id)
                return
            target_folder = test_data_folders[0]
            if not self.is_target_path_valid(target_folder):
                print(f"[see关联] 目标文件夹不在允许范围内: {target_folder}")
                self.auto_skip_no_testdata_with_see(pdf_path, original_id, see_id)
                return
            is_multi = self.multi_project_var.get()
            file_name = os.path.basename(pdf_path)
            target_path = os.path.join(target_folder, f"{settings.target_filename}.pdf")
            counter = 1
            while os.path.exists(target_path):
                target_path = os.path.join(target_folder, f"{settings.target_filename}_{counter}.pdf")
                counter += 1
            self.perform_file_operation(pdf_path, target_path, is_copy=is_multi)
            project_id_display = f"{original_id}(关联{see_id})"
            op_type = '复制(关联)' if is_multi else '移动(关联)'
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.append_record(self.saved_excel, [timestamp, file_name, project_id_display, target_folder, op_type])
            if is_multi:
                self.status_var.set(f"已复制(关联)到 {os.path.basename(target_folder)}")
                self.current_index += 1
                self.root.after(500, self.auto_run_next)
            else:
                self.pdf_files.pop(self.current_index)
                self.listbox.delete(self.current_index)
                self.current_pdf = None
                self.canvas.delete("all")
                self.original_preview_img = None
                self.folder_combo.set('')
                self.correct_entry.delete(0, tk.END)
                self.status_var.set("就绪")
                self.handle_auto_run_after_action(removed=True)
        except Exception as e:
            print(f"[错误] auto_handle_see_relation 异常: {traceback.format_exc()}")
            self.auto_skip_no_testdata_with_see(pdf_path, original_id, see_id)

    # ========== 多项目号自动处理（含对账表记录）==========
    def auto_handle_multi_projects(self):
        try:
            pdf_path = self.current_pdf
            file_name = os.path.basename(pdf_path)
            clean_project_ids = [pid.strip() for pid in self.current_all_project_ids if pid.strip()]

            all_ids_str = ", ".join(clean_project_ids)

            if len(clean_project_ids) < 2:
                if self.auto_run and not self.auto_run_paused:
                    self.root.after(0, self.auto_check_and_process)
                return

            success_count = 0
            fail_count = 0
            failed_ids = []
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            for idx, project_id in enumerate(clean_project_ids):
                current_status = "失败"
                current_remark = ""

                matches = []
                if project_id in self.search_cache:
                    matches = self.search_cache[project_id]
                else:
                    with self.thread_lock:
                        folders = self.all_project_folders.copy()
                    for folder in folders:
                        if self.folder_matches(os.path.basename(folder), project_id):
                            matches.append(folder)
                    self.search_cache[project_id] = matches
                    self.search_cache.move_to_end(project_id)
                    if len(self.search_cache) > settings.cache['search_max']:
                        self.search_cache.popitem(last=False)

                if len(matches) != 1:
                    fail_count += 1
                    fail_reason = f"{project_id} (匹配{len(matches)}个)"
                    failed_ids.append(fail_reason)
                    current_remark = f"匹配到 {len(matches)} 个文件夹，预期1个"
                else:
                    project_folder = matches[0]
                    test_data_folders = self.find_test_data_folders(project_folder)
                    if len(test_data_folders) != 1:
                        fail_count += 1
                        fail_reason = f"{project_id} (Test Data{len(test_data_folders)}个)"
                        failed_ids.append(fail_reason)
                        current_remark = f"Test Data 数量为 {len(test_data_folders)}，预期1个"
                    else:
                        target_folder = test_data_folders[0]
                        if not self.is_target_path_valid(target_folder):
                            fail_count += 1
                            fail_reason = f"{project_id} (目标文件夹不在允许范围)"
                            failed_ids.append(fail_reason)
                            current_remark = f"目标文件夹不在允许范围: {target_folder}"
                        else:
                            try:
                                target_path = os.path.join(target_folder, f"{settings.target_filename}.pdf")
                                counter = 1
                                while os.path.exists(target_path):
                                    target_path = os.path.join(target_folder, f"{settings.target_filename}_{counter}.pdf")
                                    counter += 1
                                self.perform_file_operation(pdf_path, target_path, is_copy=True)

                                current_status = "成功"
                                current_remark = f"归档至: {target_folder}"

                                self.append_record(self.saved_excel, [timestamp, file_name, project_id, target_folder, '复制(多项目)'])
                                success_count += 1
                            except Exception as e:
                                fail_count += 1
                                fail_reason = f"{project_id} (复制失败)"
                                failed_ids.append(fail_reason)
                                current_remark = f"文件复制异常: {str(e)}"

                recon_data = [
                    timestamp,
                    file_name,
                    all_ids_str,
                    project_id,
                    current_status,
                    current_remark
                ]
                self.append_record(self.reconciliation_excel, recon_data)

            print(f"[多项目处理] 完成！成功：{success_count}，失败：{fail_count}")

            if success_count > 0:
                try:
                    old_path = pdf_path
                    dir_name = os.path.dirname(old_path)
                    new_name = settings.build_filename('multi_project', original=file_name)
                    new_path = os.path.join(dir_name, new_name)
                    counter = 1
                    base, ext = os.path.splitext(new_path)
                    while os.path.exists(new_path):
                        new_path = os.path.join(dir_name, f"{base}_{counter}{ext}")
                        counter += 1
                    os.rename(old_path, new_path)

                    self.pdf_files.pop(self.current_index)
                    self.listbox.delete(self.current_index)
                    self.current_pdf = None
                    self.canvas.delete("all")
                    self.original_preview_img = None
                    self.folder_combo.set('')
                    self.correct_entry.delete(0, tk.END)
                    self.status_var.set(f"多项目处理完成：成功{success_count}个，失败{fail_count}个")
                except Exception as e:
                    print(f"[多项目处理] 重命名原文件失败：{str(e)}")
                    self.current_index += 1
            else:
                print(f"[多项目处理] 全部失败，按未找到项目处理")
                representative_id = clean_project_ids[0] if clean_project_ids else None
                self.auto_skip_no_folder_keep(pdf_path, representative_id)
                return
            self.handle_auto_run_after_action(removed=True)
        except Exception as e:
            print(f"[错误] auto_handle_multi_projects 异常: {traceback.format_exc()}")
            if self.no_intervene_var.get():
                self.auto_skip_no_folder_keep(self.current_pdf, None)
            else:
                messagebox.showerror("多项目处理异常", f"发生未预期错误：{str(e)}\n自动运行已暂停。")
                self.auto_run_paused = True

    # ========== 自动检查和处理（增加自动旋转重试，修复卡死问题）==========
    def auto_check_and_process(self):
        try:
            if not self.auto_run or self.auto_run_paused:
                return

            # 获取当前项目编号，若为空则触发自动旋转重试（只重试一次，避免多次旋转卡慢）
            has_project_id = self.correct_entry.get().strip() or self.current_all_project_ids
            if not has_project_id and self.current_pdf and settings.ocr.get('auto_rotate', True):
                # 检查是否已经旋转过（避免无限循环）
                if not getattr(self, '_auto_rotation_attempted', False):
                    self._auto_rotation_attempted = True
                    step = settings.ocr.get('rotate_step', 90) or 90
                    print(f"[自动旋转] 未识别到项目编号，尝试旋转{step}°重新识别：{os.path.basename(self.current_pdf)}")
                    self.rotation_angle = (self.rotation_angle + step) % 360
                    self.processing = True
                    thread = threading.Thread(target=self.process_pdf_thread, args=(self.current_pdf, self.rotation_angle), daemon=True)
                    thread.start()
                    return
                else:
                    # 已经尝试过旋转，仍无编号，则清除标记并进入后续处理
                    self._auto_rotation_attempted = False
                    print(f"[自动旋转] 旋转后仍无编号，进入未识别处理流程")

            # 重置旋转尝试标记（当成功识别或进入下一文件时）
            self._auto_rotation_attempted = False

            # 原有的自动归档未找到项目逻辑
            filename = os.path.basename(self.current_pdf)
            not_found_match = self.analyze_no_see_pattern.match(filename)
            if not_found_match:
                project_id = not_found_match.group(1)
                print(f"[自动归档] 检测到之前未找到的项目文件：{filename}，提取编号：{project_id}")
                matches = []
                if project_id in self.search_cache:
                    matches = self.search_cache[project_id]
                else:
                    with self.thread_lock:
                        folders = self.all_project_folders.copy()
                    for folder in folders:
                        if self.folder_matches(os.path.basename(folder), project_id):
                            matches.append(folder)
                    self.search_cache[project_id] = matches
                    self.search_cache.move_to_end(project_id)
                    if len(self.search_cache) > settings.cache['search_max']:
                        self.search_cache.popitem(last=False)

                if len(matches) == 1:
                    project_folder = matches[0]
                    test_data_folders = self.find_test_data_folders(project_folder)
                    if len(test_data_folders) == 1:
                        target_folder = test_data_folders[0]
                        if not self.is_target_path_valid(target_folder):
                            print(f"[自动归档] 目标文件夹不在允许范围: {target_folder}")
                        else:
                            target_filename = f"{settings.target_filename}.pdf"
                            target_path = os.path.join(target_folder, target_filename)
                            counter = 1
                            while os.path.exists(target_path):
                                target_path = os.path.join(target_folder, f"{settings.target_filename}_{counter}.pdf")
                                counter += 1
                            try:
                                self.perform_file_operation(self.current_pdf, target_path, is_copy=False)
                                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                self.append_record(self.saved_excel, [timestamp, filename, project_id, target_folder, '自动归档未找到项目'])
                                self.pdf_files.pop(self.current_index)
                                self.listbox.delete(self.current_index)
                                self.current_pdf = None
                                self.canvas.delete("all")
                                self.original_preview_img = None
                                self.folder_combo.set('')
                                self.correct_entry.delete(0, tk.END)
                                self.status_var.set(f"已自动归档之前未找到的项目：{filename}")
                                self.handle_auto_run_after_action(removed=True)
                                return
                            except Exception as e:
                                print(f"[错误] 自动归档失败：{e}")
                    else:
                        print(f"[自动归档] 项目 {project_id} 的 Test Data 数量异常 ({len(test_data_folders)} 个)，跳过")
                else:
                    print(f"[自动归档] 项目 {project_id} 匹配到 {len(matches)} 个文件夹，跳过")

            project_id_clean = self.correct_entry.get().strip()
            if project_id_clean and len(self.matched_folders) == 0:
                if self.current_search_retry < 2:
                    self.current_search_retry += 1
                    self.search_folders(project_id_clean, is_auto_mode=True, force_refresh=True)
                    return
                else:
                    self.auto_skip_no_folder_keep(self.current_pdf, project_id_clean)
                    return
            self.current_search_retry = 0
            if len(self.current_all_project_ids) > 1:
                self.auto_handle_multi_projects()
                return
            project_id = self.correct_entry.get().strip()
            file_name = os.path.basename(self.current_pdf)
            is_no_intervene = self.no_intervene_var.get()
            if not project_id:
                if is_no_intervene:
                    self.auto_skip_no_id(self.current_pdf)
                    return
                else:
                    self.show_pause_dialog(
                        f"文件【{file_name}】\n未识别到项目编号，即将自动重命名为“未识别_原文件名”并记录到“未找到项目”",
                        lambda: self.auto_skip_no_id(self.current_pdf)
                    )
                    return
            if len(self.matched_folders) == 0:
                if is_no_intervene:
                    self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    return
                else:
                    self.show_pause_dialog(
                        f"文件【{file_name}】\n未找到项目[{project_id}]的匹配文件夹，即将自动重命名为“未找到项目_{project_id}.pdf”并记录到“未找到项目”",
                        lambda: self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    )
                    return
            elif len(self.matched_folders) > 1:
                self.auto_copy_to_multiple_folders(self.current_pdf, project_id, self.matched_folders)
                return
            project_folder = self.matched_folders[0]
            test_data_folders = self.find_test_data_folders(project_folder)
            if len(test_data_folders) == 1:
                if not self.is_target_path_valid(test_data_folders[0]):
                    self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    return
                self.auto_move_pdf(project_folder, test_data_folders[0], self.current_pdf)
                return
            elif len(test_data_folders) > 1:
                if is_no_intervene:
                    self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    return
                else:
                    self.show_pause_dialog(
                        f"文件【{file_name}】\n项目[{project_id}]下找到{len(test_data_folders)}个Test Data文件夹，请手动选择",
                        lambda: self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    )
                    return
            see_ids = self.current_all_see_ids
            if not see_ids:
                folder_name = os.path.basename(project_folder)
                see_id_from_folder = self.extract_see_id(folder_name)
                if see_id_from_folder:
                    see_ids = [see_id_from_folder]
            if not see_ids:
                if is_no_intervene:
                    self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    return
                else:
                    self.show_pause_dialog(
                        f"文件【{file_name}】\n项目[{project_id}]下无Test Data，即将重命名为“未找到项目_{project_id}.pdf”并保留列表",
                        lambda: self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    )
                    return
            valid_see_ids = [sid for sid in see_ids if sid.strip() and sid.strip() != project_id]
            if not valid_see_ids:
                if is_no_intervene:
                    self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                else:
                    self.show_pause_dialog(
                        f"文件【{file_name}】\n识别到的关联编号和原项目重复，即将重命名为“未找到项目_{project_id}.pdf”并保留列表",
                        lambda: self.auto_skip_no_folder_keep(self.current_pdf, project_id)
                    )
                return
            all_see_matched_folders = []
            for see_id in valid_see_ids:
                if see_id in self.search_cache:
                    del self.search_cache[see_id]
                see_matches = []
                with self.thread_lock:
                    folders = self.all_project_folders.copy()
                for folder in folders:
                    if self.folder_matches(os.path.basename(folder), see_id):
                        see_matches.append(folder)
                if see_matches:
                    all_see_matched_folders.extend(see_matches)
            if not all_see_matched_folders:
                if is_no_intervene:
                    self.auto_skip_no_testdata_with_see(self.current_pdf, project_id, valid_see_ids[0])
                else:
                    self.show_pause_dialog(
                        f"文件【{file_name}】\n项目[{project_id}]下无Test Data，关联项目[{valid_see_ids[0]}]未找到匹配文件夹，即将重命名为“关联{valid_see_ids[0]}_未找到项目_{project_id}.pdf”并保留列表",
                        lambda: self.auto_skip_no_testdata_with_see(self.current_pdf, project_id, valid_see_ids[0])
                    )
                return
            self.auto_copy_to_multiple_folders(
                self.current_pdf,
                f"{project_id}(关联{','.join(valid_see_ids)})",
                all_see_matched_folders
            )
        except Exception as e:
            print(f"[错误] auto_check_and_process 异常: {traceback.format_exc()}")
            if self.no_intervene_var.get():
                self.auto_skip_no_folder_keep(self.current_pdf, None)
            else:
                messagebox.showerror("自动处理异常", f"发生未预期错误：{str(e)}\n自动运行已暂停。")
                self.auto_run_paused = True

    def auto_move_pdf(self, project_folder, target_folder, pdf_path):
        try:
            is_multi = self.multi_project_var.get()
            project_id = self.correct_entry.get().strip()
            if not self.is_target_path_valid(target_folder):
                raise Exception("目标文件夹不在允许的根目录范围内")
            target_path = os.path.join(target_folder, f"{settings.target_filename}.pdf")
            counter = 1
            while os.path.exists(target_path):
                target_path = os.path.join(target_folder, f"{settings.target_filename}_{counter}.pdf")
                counter += 1
            self.perform_file_operation(pdf_path, target_path, is_copy=is_multi)
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.append_record(self.saved_excel, [timestamp, os.path.basename(pdf_path),
                                                   project_id, target_folder, '复制' if is_multi else '移动'])
            if is_multi:
                self.status_var.set(f"已自动复制到 {os.path.basename(target_folder)}")
                self.current_index += 1
                self.root.after(500, self.auto_run_next)
            else:
                self.pdf_files.pop(self.current_index)
                self.listbox.delete(self.current_index)
                self.current_pdf = None
                self.canvas.delete("all")
                self.original_preview_img = None
                self.folder_combo.set('')
                self.correct_entry.delete(0, tk.END)
                self.status_var.set("就绪")
                self.handle_auto_run_after_action(removed=True)
        except Exception as e:
            self.status_var.set(f"自动处理失败: {str(e)}")
            if self.no_intervene_var.get():
                self.auto_skip_no_folder_keep(pdf_path, project_id if 'project_id' in locals() else None)
            else:
                self.auto_run_paused = True
                messagebox.showerror("自动处理失败", str(e))

    def handle_auto_run_after_action(self, removed):
        if not self.auto_run or self.auto_run_paused:
            return
        if removed:
            pass
        else:
            self.current_index += 1
        self.root.after(500, self.auto_run_next)

    # ========== 图像预处理 ==========
    def preprocess_image(self, img):
        if img.mode != 'L':
            img = img.convert('L')
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.5)
        return img

    # ========== 不识别模式：仅显示预览（支持旋转角度）==========
    def generate_thumbnail_only(self, pdf_path, angle=0):
        try:
            safe_path, is_temp = self.get_safe_pdf_path(pdf_path)
            try:
                images = self.run_with_timeout(
                    convert_from_path,
                    args=(safe_path,),
                    kwargs={
                        'dpi': settings.ocr['dpi'],
                        'poppler_path': POPPLER_PATH,
                        'first_page': 1,
                        'last_page': 1,
                        'strict': False,
                        'use_cropbox': True,
                    },
                    timeout=settings.ocr['preview_timeout_sec']
                )
            finally:
                if is_temp and os.path.exists(safe_path):
                    try:
                        os.remove(safe_path)
                    except:
                        pass
            if not images:
                raise Exception("PDF无有效页面")

            img = images[0]
            if angle != 0:
                img = img.rotate(angle, expand=True)

            self.original_preview_img = img.copy()
            self.update_canvas_image()

            self.current_text = ""
            self.current_all_project_ids = []
            self.current_all_see_ids = []
            self.matched_folders = []

            self.status_var.set(f"不识别模式：已旋转{angle}°，仅更新预览，未进行OCR识别")
        except Exception as e:
            self.status_var.set(f"预览旋转失败: {str(e)}")
            self.processing = False

    def show_ocr_text(self):
        if self.no_ocr_var.get():
            messagebox.showinfo("提示", "当前处于“不识别模式”，未进行OCR识别，无文本可显示。")
            return
        if not self.current_text or self.current_text.strip() == "":
            messagebox.showinfo("提示", "当前没有可显示的识别文本\n请先选择一个PDF文件，等待OCR识别完成后再试。")
            return
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Toplevel) and "OCR识别文本" in widget.title():
                widget.lift()
                widget.focus_force()
                return
        text_window = tk.Toplevel(self.root)
        text_window.title(f"OCR识别文本 - {os.path.basename(self.current_pdf)}")
        text_window.geometry("800x600")
        text_window.minsize(500, 350)
        text_window.transient(self.root)
        text_window.grab_set()
        frame = tk.Frame(text_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        text_widget = tk.Text(frame, wrap=tk.WORD, font=("微软雅黑", 10), state=tk.NORMAL)
        y_scroll = tk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        x_scroll = tk.Scrollbar(frame, orient=tk.HORIZONTAL, command=text_widget.xview)
        text_widget.config(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        text_widget.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        text_widget.delete(1.0, tk.END)
        text_widget.insert(tk.END, self.current_text)
        text_widget.config(state=tk.DISABLED)
        btn_frame = tk.Frame(text_window)
        btn_frame.pack(fill=tk.X, pady=8, padx=8)
        def copy_full_text():
            self.root.clipboard_clear()
            self.root.clipboard_append(self.current_text)
            messagebox.showinfo("复制成功", "识别文本已全部复制到剪贴板", parent=text_window)
        tk.Button(btn_frame, text="复制全部文本", command=copy_full_text, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="关闭窗口", command=text_window.destroy, width=12).pack(side=tk.RIGHT, padx=5)
        text_window.update_idletasks()
        width = text_window.winfo_width()
        height = text_window.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        text_window.geometry(f"{width}x{height}+{x}+{y}")

    def show_fullsize_preview(self):
        if not self.current_pdf:
            messagebox.showinfo("提示", "请先选择一个PDF文件")
            return
        try:
            safe_path, is_temp = self.get_safe_pdf_path(self.current_pdf)
            try:
                images = self.run_with_timeout(
                    convert_from_path,
                    args=(safe_path,),
                    kwargs={
                        'dpi': 150,
                        'poppler_path': POPPLER_PATH,
                        'first_page': 1,
                        'last_page': 1,
                        'strict': False,
                        'use_cropbox': True,
                    },
                    timeout=settings.ocr['preview_timeout_sec']
                )
            finally:
                if is_temp and os.path.exists(safe_path):
                    try:
                        os.remove(safe_path)
                    except:
                        pass
            if not images:
                raise Exception("无法获取PDF页面")
            img = images[0]
            if self.rotation_angle != 0:
                img = img.rotate(self.rotation_angle, expand=True)
            full_win = tk.Toplevel(self.root)
            full_win.title(f"全尺寸预览 - {os.path.basename(self.current_pdf)}")
            full_win.geometry("800x600")
            full_win.minsize(400, 300)
            canvas = tk.Canvas(full_win, bg='gray')
            v_scroll = tk.Scrollbar(full_win, orient=tk.VERTICAL, command=canvas.yview)
            h_scroll = tk.Scrollbar(full_win, orient=tk.HORIZONTAL, command=canvas.xview)
            canvas.config(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
            canvas.grid(row=0, column=0, sticky="nsew")
            v_scroll.grid(row=0, column=1, sticky="ns")
            h_scroll.grid(row=1, column=0, sticky="ew")
            full_win.grid_rowconfigure(0, weight=1)
            full_win.grid_columnconfigure(0, weight=1)
            photo = ImageTk.PhotoImage(img)
            canvas.create_image(0, 0, anchor=tk.NW, image=photo)
            canvas.config(scrollregion=canvas.bbox("all"))
            canvas.image = photo
            tk.Button(full_win, text="关闭", command=full_win.destroy, width=10).grid(row=2, column=0, pady=5)
        except Exception as e:
            messagebox.showerror("预览失败", f"无法加载全尺寸图片：{str(e)}")

    # ========== 界面创建 ==========
    def create_widgets(self):
        # 主容器
        main_container = tk.Frame(self.root)
        main_container.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        main_container.grid_rowconfigure(3, weight=1)
        main_container.grid_columnconfigure(0, weight=1)

        # 顶部：源文件夹
        top_frame = tk.LabelFrame(main_container, text="文件源设置", padx=8, pady=5)
        top_frame.grid(row=0, column=0, sticky="ew", pady=3)
        top_frame.grid_columnconfigure(1, weight=10)
        tk.Label(top_frame, text="源文件夹:", font=("微软雅黑", 9)).grid(row=0, column=0, sticky="w", padx=3)
        self.entry_source = tk.Entry(top_frame, textvariable=self.source_folder, font=("微软雅黑", 9))
        self.entry_source.grid(row=0, column=1, sticky="ew", padx=5)
        tk.Button(top_frame, text="浏览", command=self.select_source_folder, width=8).grid(row=0, column=2, padx=2)
        tk.Button(top_frame, text="刷新列表", command=self.refresh_pdf_list, width=8).grid(row=0, column=3, padx=2)
        tk.Button(top_frame, text="⚙ 设置", command=self.open_settings, width=8).grid(row=0, column=4, padx=2)
        tk.Button(top_frame, text="❓ 帮助", command=self.show_help, width=8).grid(row=0, column=5, padx=2)

        # 年份选择栏
        year_frame = tk.Frame(main_container)
        year_frame.grid(row=1, column=0, sticky="ew", pady=2)
        tk.Label(year_frame, text="分拣年份:").pack(side=tk.LEFT, padx=3)
        self.year_combo = ttk.Combobox(year_frame, textvariable=self.selected_year_var, values=self.year_options, state="normal", width=10)
        self.year_combo.pack(side=tk.LEFT, padx=5)
        self.year_combo.bind("<<ComboboxSelected>>", self.on_year_changed)
        tk.Label(year_frame, text="(输入年份如25 → 点击重建索引)", fg="gray", font=("微软雅黑", 8)).pack(side=tk.LEFT, padx=5)

        # 目标范围框 - 紧凑布局
        range_frame = tk.LabelFrame(main_container, text="目标范围", padx=3, pady=2)
        range_frame.grid(row=2, column=0, sticky="ew", pady=2)
        range_frame.grid_columnconfigure(0, weight=1)
        
        # 列表框架（放在上方）
        list_frame = tk.Frame(range_frame)
        list_frame.grid(row=0, column=0, sticky="nsew")
        list_frame.grid_columnconfigure(0, weight=1)
        
        # 列表内容（去掉标题，减少高度）
        self.roots_listbox = tk.Listbox(list_frame, height=3, font=("微软雅黑", 8))
        self.roots_listbox.grid(row=0, column=0, sticky="nsew")
        scrollbar = tk.Scrollbar(list_frame, width=12)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.roots_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.roots_listbox.yview)
        
        # 按钮框架（横向排列，放在列表下方）
        btn_range_frame = tk.Frame(range_frame)
        btn_range_frame.grid(row=1, column=0, sticky="ew", pady=2)
        
        # 左侧按钮组
        left_btn_frame = tk.Frame(btn_range_frame)
        left_btn_frame.pack(side=tk.LEFT)
        tk.Button(left_btn_frame, text="添加根目录", command=self.add_target_root, width=10).pack(side=tk.LEFT, padx=2)
        tk.Button(left_btn_frame, text="删除选中", command=self.remove_target_root, width=10).pack(side=tk.LEFT, padx=2)
        tk.Button(left_btn_frame, text="编辑深度", command=self.edit_root_depth, width=10).pack(side=tk.LEFT, padx=2)
        
        # 右侧按钮组
        right_btn_frame = tk.Frame(btn_range_frame)
        right_btn_frame.pack(side=tk.RIGHT)
        tk.Button(right_btn_frame, text="重建索引", command=self.refresh_index, width=10).pack(side=tk.LEFT, padx=2)
        tk.Button(right_btn_frame, text="刷新选中", command=self.refresh_selected_root, width=10).pack(side=tk.LEFT, padx=2)
        
        for root_item in self.target_roots:
            root_path = root_item['path']
            depth = root_item.get('depth', settings.cache['search_depth'])
            display_text = f"{root_path:.<60}  [{depth}层]"
            self.roots_listbox.insert(tk.END, display_text)

        # 主内容区：左侧文件列表 + 右侧预览
        main_frame = tk.Frame(main_container)
        main_frame.grid(row=3, column=0, sticky="nsew", pady=3)
        main_frame.grid_columnconfigure(0, weight=0, minsize=220)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)

        left_frame = tk.LabelFrame(main_frame, text="PDF文件列表", padx=5, pady=5)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0,3))
        left_frame.grid_rowconfigure(0, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)
        self.listbox = tk.Listbox(left_frame, font=("微软雅黑", 9))
        self.listbox.grid(row=0, column=0, sticky="nsew")
        self.listbox.bind('<<ListboxSelect>>', self.on_select_pdf)

        right_frame = tk.Frame(main_frame)
        right_frame.grid(row=0, column=1, sticky="nsew")
        right_frame.grid_rowconfigure(1, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        preview_control_frame = tk.Frame(right_frame)
        preview_control_frame.grid(row=0, column=0, sticky="ew", pady=2)
        tk.Label(preview_control_frame, text="PDF预览（可缩放拖动）").pack(side=tk.LEFT)
        tk.Button(preview_control_frame, text="放大", command=self.zoom_in, width=6).pack(side=tk.LEFT, padx=3)
        tk.Button(preview_control_frame, text="缩小", command=self.zoom_out, width=6).pack(side=tk.LEFT, padx=3)
        tk.Button(preview_control_frame, text="全尺寸预览", command=self.show_fullsize_preview, width=10, bg='lightblue').pack(side=tk.LEFT, padx=5)
        tk.Button(preview_control_frame, text="查看识别文本", command=self.show_ocr_text, width=12, bg='lightyellow').pack(side=tk.LEFT, padx=5)

        canvas_container = tk.Frame(right_frame)
        canvas_container.grid(row=1, column=0, sticky="nsew", pady=3)
        canvas_container.grid_rowconfigure(0, weight=1)
        canvas_container.grid_columnconfigure(0, weight=1)
        self.canvas = tk.Canvas(canvas_container, bg='gray', highlightthickness=0)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        v_scroll = tk.Scrollbar(canvas_container, orient=tk.VERTICAL, command=self.canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll = tk.Scrollbar(canvas_container, orient=tk.HORIZONTAL, command=self.canvas.xview)
        h_scroll.grid(row=1, column=0, sticky="ew")
        self.canvas.config(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        self.canvas.bind("<ButtonPress-1>", self.on_drag_start)
        self.canvas.bind("<B1-Motion>", self.on_drag_move)
        self.canvas.bind("<Enter>", lambda e: self.canvas.config(cursor="hand2"))
        self.canvas.bind("<Leave>", lambda e: self.canvas.config(cursor=""))
        self.canvas.bind("<Control-MouseWheel>", self.on_ctrl_mousewheel)
        self.canvas.bind("<Control-Button-4>", self.on_ctrl_mousewheel)
        self.canvas.bind("<Control-Button-5>", self.on_ctrl_mousewheel)

        # 底部操作栏
        bottom_frame = tk.LabelFrame(main_container, text="操作区", padx=5, pady=5)
        bottom_frame.grid(row=4, column=0, sticky="ew", pady=3)
        bottom_frame.grid_columnconfigure(0, weight=1)

        match_frame = tk.Frame(bottom_frame)
        match_frame.grid(row=0, column=0, sticky="ew", pady=2)
        match_frame.grid_columnconfigure(1, weight=1)
        tk.Label(match_frame, text="目标文件夹:").pack(side=tk.LEFT)
        self.folder_combo = ttk.Combobox(match_frame, state='normal', font=("微软雅黑", 8))
        self.folder_combo.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        tk.Button(match_frame, text="浏览", command=self.browse_target_folder, width=8).pack(side=tk.LEFT)

        correct_frame = tk.Frame(bottom_frame)
        correct_frame.grid(row=1, column=0, sticky="ew", pady=2)
        correct_frame.grid_columnconfigure(1, weight=1)
        tk.Label(correct_frame, text="手动修正编号:").pack(side=tk.LEFT)
        self.correct_entry = tk.Entry(correct_frame, font=("微软雅黑", 9))
        self.correct_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        tk.Button(correct_frame, text="重新匹配", command=self.rematch_with_corrected, width=10).pack(side=tk.LEFT)

        mode_frame = tk.Frame(bottom_frame)
        mode_frame.grid(row=2, column=0, sticky="ew", pady=2)
        self.multi_check = tk.Checkbutton(mode_frame, text="多项目复制模式", variable=self.multi_project_var)
        self.multi_check.pack(side=tk.LEFT, padx=5)
        self.no_intervene_check = tk.Checkbutton(mode_frame, text="🚀 全自动不干预模式", variable=self.no_intervene_var)
        self.no_intervene_check.pack(side=tk.LEFT, padx=10)
        self.no_ocr_check = tk.Checkbutton(mode_frame, text="😴 不识别模式（跳过OCR，仅显示图片）", variable=self.no_ocr_var)
        self.no_ocr_check.pack(side=tk.LEFT, padx=10)

        btn_frame = tk.Frame(bottom_frame)
        btn_frame.grid(row=3, column=0, sticky="ew", pady=3)
        btn_frame.grid_columnconfigure(0, weight=1)

        btn_container = tk.Frame(btn_frame)
        btn_container.pack(anchor="center")

        self.btn_confirm = tk.Button(btn_container, text="确认移动", command=self.move_pdf, bg='lightblue', width=12)
        self.btn_confirm.grid(row=0, column=0, padx=3)
        self.btn_skip = tk.Button(btn_container, text="跳过", command=self.skip_pdf, width=12)
        self.btn_skip.grid(row=0, column=1, padx=3)
        self.btn_not_found = tk.Button(btn_container, text="未找到项目", command=self.not_found, bg='orange', width=12)
        self.btn_not_found.grid(row=0, column=2, padx=3)
        self.btn_rotate = tk.Button(btn_container, text="旋转90°", command=self.rotate_image, bg='lightgreen', width=12)
        self.btn_rotate.grid(row=0, column=3, padx=3)
        self.btn_auto_run = tk.Button(btn_container, text="自动运行", command=self.start_auto_run, bg='lightgreen', width=12)
        self.btn_auto_run.grid(row=0, column=4, padx=3)
        self.btn_continue = tk.Button(btn_container, text="继续自动", command=self.continue_auto_run, bg='lightyellow', width=12)
        self.btn_continue.grid(row=0, column=5, padx=3)
        self.btn_exit = tk.Button(btn_container, text="退出", command=self.on_closing, width=12)
        self.btn_exit.grid(row=0, column=6, padx=3)

        status_frame = tk.Frame(bottom_frame)
        status_frame.grid(row=4, column=0, sticky="ew", pady=2)
        status_label = tk.Label(status_frame, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W, font=("微软雅黑", 9))
        status_label.pack(fill=tk.X)

        self.progress_label = tk.Label(
            main_container,
            textvariable=self.index_progress_var,
            fg="darkblue",
            font=("微软雅黑", 8)
        )
        self.progress_label.grid(row=5, column=0, sticky="se", padx=5, pady=2)

    # ========== 主要功能 ==========
    def select_source_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.source_folder.set(folder)
            self.update_record_files()
            self.refresh_pdf_list()

    def refresh_pdf_list(self):
        folder = self.source_folder.get()
        if not folder or not os.path.isdir(folder):
            return
        self.pdf_files = []
        self.listbox.delete(0, tk.END)
        for f in os.listdir(folder):
            if f.lower().endswith('.pdf'):
                full_path = os.path.join(folder, f)
                self.pdf_files.append(full_path)
                self.listbox.insert(tk.END, f)
        self.status_var.set(f"找到 {len(self.pdf_files)} 个PDF文件")

    def get_ocr_cache_key(self, pdf_path):
        """生成带修改时间的缓存key，文件修改后自动失效"""
        mtime = os.path.getmtime(pdf_path)
        return f"{pdf_path}_{mtime}"

    def on_select_pdf(self, event):
        if self.auto_run and self.processing:
            return
        if self.processing:
            messagebox.showinfo("提示", "正在处理上一个文件，请稍候")
            return
        selection = self.listbox.curselection()
        if not selection:
            return
        index = selection[0]
        self.current_index = index
        pdf_path = self.pdf_files[index]

        self.rotation_angle = 0
        self.scale_factor = 1.0
        self.current_search_retry = 0
        self._auto_rotation_attempted = False  # 重置旋转标记

        if self.no_ocr_var.get():
            self.current_pdf = pdf_path
            self.rotation_angle = 0
            self.generate_thumbnail_only(pdf_path)
            self.processing = False
            return

        see_id_from_filename = self.extract_see_from_filename(os.path.basename(pdf_path))
        if see_id_from_filename:
            self.correct_entry.delete(0, tk.END)
            self.correct_entry.insert(0, see_id_from_filename)

        # 检查缓存（带文件修改时间）
        cache_key = self.get_ocr_cache_key(pdf_path)
        if cache_key in self.ocr_cache:
            self.current_pdf = pdf_path
            self.current_text = self.ocr_cache[cache_key]
            self.generate_thumbnail_and_ocr(pdf_path, angle=0)
            self.status_var.set("已从缓存加载")
            return

        self.current_pdf = pdf_path
        self.status_var.set(f"正在识别: {os.path.basename(pdf_path)} ...")
        self.root.update()

        self.canvas.delete("all")
        self.correct_entry.delete(0, tk.END)

        self.processing = True
        thread = threading.Thread(target=self.process_pdf_thread, args=(pdf_path, 0), daemon=True)
        thread.start()

    def generate_thumbnail_and_ocr(self, pdf_path, angle):
        try:
            safe_path, is_temp = self.get_safe_pdf_path(pdf_path)
            try:
                images = self.run_with_timeout(
                    convert_from_path,
                    args=(safe_path,),
                    kwargs={
                        'dpi': settings.ocr['dpi'],
                        'poppler_path': POPPLER_PATH,
                        'first_page': 1,
                        'last_page': 1,
                        'strict': False,
                        'use_cropbox': True,
                    },
                    timeout=settings.ocr['preview_timeout_sec']
                )
            finally:
                if is_temp and os.path.exists(safe_path):
                    try:
                        os.remove(safe_path)
                    except:
                        pass

            if not images:
                raise Exception("PDF无有效页面，无法解析")

            img = images[0]
            if angle != 0:
                img = img.rotate(angle, expand=True)
            self.original_preview_img = img.copy()
            self.update_canvas_image()

            cache_key = self.get_ocr_cache_key(pdf_path)
            if cache_key in self.ocr_cache and angle == 0:
                text = self.ocr_cache[cache_key]
            else:
                text = pytesseract.image_to_string(img, lang=settings.ocr_lang_string(), config=settings.ocr['config'])
                if angle == 0:
                    self.ocr_cache[cache_key] = text
                    self.ocr_cache.move_to_end(cache_key)
                    if len(self.ocr_cache) > settings.cache['ocr_max']:
                        self.ocr_cache.popitem(last=False)

            self.current_text = text
            print(f"[OCR调试] 识别完成，文本长度：{len(self.current_text)} 字符")

            self.current_all_project_ids = self.extract_all_project_ids(text)
            self.current_all_see_ids = self.extract_all_see_ids(text)

            final_project_id = None
            if self.current_all_project_ids:
                final_project_id = self.current_all_project_ids[0]
            elif self.current_all_see_ids:
                final_project_id = self.current_all_see_ids[0]

            if not final_project_id:
                see_id_from_filename = self.extract_see_from_filename(os.path.basename(pdf_path))
                if see_id_from_filename:
                    final_project_id = see_id_from_filename

            if final_project_id:
                self.correct_entry.delete(0, tk.END)
                self.correct_entry.insert(0, final_project_id)
                self.search_folders(final_project_id, is_auto_mode=self.auto_run)
            else:
                self.status_var.set("未识别到项目编号和关联编号，请手动输入")
                if not self.auto_run:
                    messagebox.showwarning("未识别", "未能从PDF中识别出项目编号和关联编号，请手动输入后点击'重新匹配'")
                if self.auto_run and not self.auto_run_paused:
                    self.root.after(0, self.auto_check_and_process)

        except TimeoutError as e:
            error_msg = f"生成预览超时：{str(e)}"
            print(f"[预览超时] 文件：{os.path.basename(pdf_path)} | {error_msg}")
            self.processing = False
            if self.auto_run:
                self.on_ocr_error(error_msg, pdf_path)
            else:
                messagebox.showerror("处理失败", error_msg)
        except Exception as e:
            error_msg = f"生成预览或OCR时出错：{str(e)}"
            print(f"[预览错误] 文件：{os.path.basename(pdf_path)} | {error_msg}")
            self.processing = False
            if self.auto_run:
                self.on_ocr_error(error_msg, pdf_path)
            else:
                messagebox.showerror("处理失败", error_msg)

    def process_pdf_thread(self, pdf_path, angle):
        try:
            safe_path, is_temp = self.get_safe_pdf_path(pdf_path)
            try:
                print(f"[OCR调试] 开始转换PDF：{os.path.basename(pdf_path)}")
                images = self.run_with_timeout(
                    convert_from_path,
                    args=(safe_path,),
                    kwargs={
                        'dpi': settings.ocr['dpi'],
                        'poppler_path': POPPLER_PATH,
                        'first_page': 1,
                        'last_page': settings.ocr['max_pages'],
                        'thread_count': 1,
                        'use_cropbox': True,
                        'strict': False
                    },
                    timeout=settings.ocr['convert_timeout_sec']
                )
            finally:
                if is_temp and os.path.exists(safe_path):
                    try:
                        os.remove(safe_path)
                    except:
                        pass

            if not images:
                self.root.after(0, self.on_ocr_error, "PDF无有效页面，无法解析", pdf_path)
                return

            rotated_images = []
            for i, img in enumerate(images):
                if i >= settings.ocr['max_pages']:
                    break
                if angle != 0:
                    img = img.rotate(angle, expand=True)
                rotated_images.append(img)

            full_text = ""
            print(f"[OCR调试] 开始OCR识别（角度：{angle}°）：{os.path.basename(pdf_path)}")
            for i, img in enumerate(rotated_images):
                page_text = self.run_with_timeout(
                    pytesseract.image_to_string,
                    args=(img,),
                    kwargs={'lang': settings.ocr_lang_string(), 'config': settings.ocr['config']},
                    timeout=settings.ocr['page_timeout_sec']
                )
                full_text += f"--- 第{i+1}页 ---\n{page_text}\n"

            if angle == 0:
                cache_key = self.get_ocr_cache_key(pdf_path)
                self.ocr_cache[cache_key] = full_text
                self.ocr_cache.move_to_end(cache_key)
                if len(self.ocr_cache) > settings.cache['ocr_max']:
                    self.ocr_cache.popitem(last=False)

            self.root.after(0, self.on_ocr_done, full_text, pdf_path)

        except TimeoutError as e:
            print(f"[OCR超时] 文件：{os.path.basename(pdf_path)} | 错误：{str(e)}")
            self.root.after(0, self.on_ocr_error, f"文件处理超时，已跳过", pdf_path)
        except Exception as e:
            print(f"[OCR异常] 文件：{os.path.basename(pdf_path)} | 错误：{traceback.format_exc()}")
            self.root.after(0, self.on_ocr_error, str(e), pdf_path)

    def on_ocr_done(self, text, pdf_path):
        self.current_text = text
        self.generate_thumbnail_and_ocr(pdf_path, self.rotation_angle)
        self.processing = False

    def on_ocr_error(self, error_msg, pdf_path):
        self.status_var.set(f"OCR出错: {error_msg}")
        print(f"[错误] OCR出错: {error_msg}，文件：{os.path.basename(pdf_path)}")
        self.processing = False
        if self.auto_run:
            if self.no_intervene_var.get():
                print(f"[不干预模式] OCR失败，自动跳过文件：{os.path.basename(pdf_path)}")
                if pdf_path == self.current_pdf:
                    try:
                        idx = self.pdf_files.index(pdf_path)
                        self.pdf_files.pop(idx)
                        self.listbox.delete(idx)
                    except:
                        pass
                self.handle_auto_run_after_action(removed=True)
            else:
                self.auto_run_paused = True
                messagebox.showerror("OCR错误", f"处理PDF【{os.path.basename(pdf_path)}】时出错:\n{error_msg}\n自动运行已暂停")
                self.status_var.set("OCR识别失败，自动运行已暂停，请手动处理")
        else:
            messagebox.showerror("OCR错误", f"处理PDF【{os.path.basename(pdf_path)}】时出错:\n{error_msg}")

    def extract_project_id(self, text):
        match = self.project_regex.search(text)
        return match.group(0).strip() if match else None

    def extract_see_id(self, folder_name):
        match = self.see_regex.search(folder_name)
        return match.group(1).strip() if match else None

    def folder_matches(self, folder_name, project_id):
        if settings.folder_match.get('prefix_match', True) and folder_name.lower().startswith(project_id.lower()):
            return True
        if settings.folder_match.get('word_boundary_match', True):
            pattern = r'(?:^|\W)' + re.escape(project_id) + r'(?:$|\W)'
            return re.search(pattern, folder_name, re.IGNORECASE) is not None
        return False

    def search_folders(self, project_id, is_auto_mode=False, force_refresh=False):
        project_id_clean = project_id.strip()
        if not project_id_clean:
            self.matched_folders = []
            self.root.after(0, lambda: self._update_folder_combo([]))
            return
        if self.search_timer:
            self.root.after_cancel(self.search_timer)
            self.search_timer = None
        if force_refresh and project_id_clean in self.search_cache:
            del self.search_cache[project_id_clean]
            print(f"[搜索调试] 强制刷新，已清除缓存：{project_id_clean}")
        self.matched_folders = []
        self.root.after(0, lambda: self._update_folder_combo([]))
        self.status_var.set(f"正在搜索项目: {project_id_clean}...")
        delay_ms = 0 if is_auto_mode else 300
        self.search_timer = self.root.after(delay_ms, lambda: self._do_search_folders(project_id_clean))

    def _do_search_folders(self, project_id_clean):
        try:
            print(f"[搜索调试] 执行搜索 | 干净项目号：{repr(project_id_clean)} | 自动模式：{self.auto_run}")
            if self.indexing:
                self.root.after(0, lambda: self.status_var.set("索引正在构建中，搜索结果可能不完整"))
            if not self.all_project_folders:
                self.root.after(0, lambda: self.status_var.set("项目文件夹索引为空，请添加目标范围并点击'重建索引'"))
                self.matched_folders = []
                self.root.after(0, lambda: self._update_folder_combo([]))
                return
            if project_id_clean in self.search_cache:
                matches = self.search_cache[project_id_clean]
                self.matched_folders = matches
                self.root.after(0, lambda: self._update_folder_combo(matches))
                print(f"[搜索调试] 缓存命中 | 项目号：{project_id_clean} | 匹配数量：{len(matches)}")
                if self.auto_run and not self.auto_run_paused:
                    self.root.after(0, self.auto_check_and_process)
                return
            matches = []
            with self.thread_lock:
                folders = self.all_project_folders.copy()
            for folder in folders:
                if self.folder_matches(os.path.basename(folder), project_id_clean):
                    matches.append(folder)
            self.matched_folders = matches
            self.search_cache[project_id_clean] = matches
            self.search_cache.move_to_end(project_id_clean)
            if len(self.search_cache) > settings.cache['search_max']:
                self.search_cache.popitem(last=False)
            self.root.after(0, lambda: self._update_folder_combo(matches))
            print(f"[搜索调试] 实时搜索完成 | 项目号：{project_id_clean} | 匹配数量：{len(matches)}")
            if self.auto_run and not self.auto_run_paused:
                self.root.after(0, self.auto_check_and_process)
        except Exception as e:
            print(f"[错误] _do_search_folders 异常: {traceback.format_exc()}")
            self.root.after(0, lambda: self.status_var.set("搜索出错，请重试"))
            if self.auto_run and not self.auto_run_paused:
                self.root.after(0, self.auto_check_and_process)

    def _update_folder_combo(self, matches):
        display_list = matches.copy()
        final_target = None
        if matches:
            primary_project = matches[0]
            test_data_folders = self.find_test_data_folders(primary_project)
            if test_data_folders:
                # 只显示合法的 Test Data 文件夹
                valid_test_data = [f for f in test_data_folders if self.is_target_path_valid(f)]
                if valid_test_data:
                    display_list = matches + valid_test_data
                    final_target = primary_project
                    self.status_var.set(f"找到 {len(matches)} 个项目文件夹，及其 {len(valid_test_data)} 个 Test Data 子文件夹")
                else:
                    final_target = primary_project
                    self.status_var.set(f"找到 {len(matches)} 个项目文件夹，但未找到可用的 Test Data")
            else:
                final_target = primary_project
                self.status_var.set(f"找到 {len(matches)} 个项目文件夹，但未找到 Test Data")
            self.folder_combo['values'] = display_list
            self.folder_combo.set(final_target)
        else:
            self.folder_combo['values'] = []
            self.folder_combo.set('')
            self.status_var.set("未找到匹配的项目文件夹" + ("（部分索引）" if self.indexing else ""))

    def rematch_with_corrected(self):
        corrected = self.correct_entry.get().strip()
        if not corrected:
            messagebox.showwarning("输入错误", "请输入项目编号")
            return
        self.search_folders(corrected, is_auto_mode=False, force_refresh=True)

    def rotate_image(self):
        if not self.current_pdf:
            messagebox.showwarning("提示", "请先选择一个PDF文件")
            return
        if self.processing:
            messagebox.showinfo("提示", "正在处理中，请稍候")
            return

        self.rotation_angle = (self.rotation_angle + 90) % 360
        self.status_var.set(f"正在旋转 {self.rotation_angle}°...")
        self.root.update()

        if self.no_ocr_var.get():
            self.processing = True
            self.generate_thumbnail_only(self.current_pdf, angle=self.rotation_angle)
            self.processing = False
            return

        self.processing = True
        thread = threading.Thread(target=self.process_pdf_thread, args=(self.current_pdf, self.rotation_angle), daemon=True)
        thread.start()

    def find_test_data_folders(self, project_folder):
        test_data_folders = []
        # 目标子目录名匹配：不区分大小写，忽略空格/下划线/横线等分隔符差异
        try:
            if not os.path.exists(project_folder):
                return []
            items = os.listdir(project_folder)
            for item in items:
                item_path = os.path.join(project_folder, item)
                if os.path.isdir(item_path) and settings.subfolder_matches(item):
                    test_data_folders.append(item_path)
        except Exception as e:
            print(f"[调试] 读取项目文件夹失败 [{project_folder}]: {e}")
        return test_data_folders

    # ========== 手动操作函数 ==========
    @button_guard
    def move_pdf(self):
        self.cancel_auto_timer()
        self.close_pause_dialog()
        if self.auto_run and self.auto_run_paused:
            self.auto_run_paused = False

        if not self.current_pdf:
            return
        target_path_input = self.folder_combo.get().strip()
        if not target_path_input:
            messagebox.showerror("错误", "请选择或输入目标文件夹")
            return

        # 解析目标文件夹（子目录名判断：不区分大小写，忽略空格/分隔符差异）
        if settings.subfolder_matches(os.path.basename(target_path_input)):
            target_folder = target_path_input
            project_folder = os.path.dirname(target_folder)
        else:
            project_folder = target_path_input
            test_data_folders = self.find_test_data_folders(project_folder)
            if not test_data_folders:
                messagebox.showerror("错误", f"项目文件夹 {os.path.basename(project_folder)} 下无「{settings.subfolder_name}」文件夹")
                return
            if len(test_data_folders) == 1:
                target_folder = test_data_folders[0]
            else:
                folder_names = [os.path.basename(f) for f in test_data_folders]
                selected = simpledialog.askstring(f"选择 {settings.subfolder_name} 文件夹",
                                                  f"找到多个 {settings.subfolder_name} 子文件夹，请选择其中一个：\n"
                                                  f"{', '.join(folder_names)}\n"
                                                  f"输入完整文件夹名（如 '01 {settings.subfolder_name}'）：")
                if not selected or selected not in folder_names:
                    messagebox.showerror("错误", "输入无效，操作取消")
                    return
                target_folder = os.path.join(project_folder, selected)

        if not os.path.exists(target_folder):
            messagebox.showerror("错误", f"目标文件夹不存在：{target_folder}")
            return

        # 安全校验
        if not self.is_target_path_valid(target_folder):
            messagebox.showerror("错误", "目标文件夹不在允许的根目录范围内，禁止操作！")
            return

        is_multi = self.multi_project_var.get()
        project_id = self.correct_entry.get().strip()
        if not project_id:
            project_id = self.extract_project_id(self.current_text) or '未知编号'

        target_path = os.path.join(target_folder, f"{settings.target_filename}.pdf")
        counter = 1
        while os.path.exists(target_path):
            target_path = os.path.join(target_folder, f"{settings.target_filename}_{counter}.pdf")
            counter += 1

        removed = not is_multi
        op_type = '复制' if is_multi else '移动'

        try:
            self.perform_file_operation(self.current_pdf, target_path, is_copy=is_multi)
            messagebox.showinfo("成功", f"文件已{op_type}到:\n{target_path}")
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.append_record(self.saved_excel, [timestamp, os.path.basename(self.current_pdf),
                                                   project_id, target_folder, op_type])
            if is_multi:
                self.status_var.set(f"已复制到 {os.path.basename(target_folder)}")
            else:
                self.pdf_files.pop(self.current_index)
                self.listbox.delete(self.current_index)
                self.current_pdf = None
                self.canvas.delete("all")
                self.original_preview_img = None
                self.folder_combo.set('')
                self.correct_entry.delete(0, tk.END)
                self.status_var.set("就绪")
            if self.auto_run:
                self.handle_auto_run_after_action(removed)
        except Exception as e:
            messagebox.showerror("操作失败", str(e))

    @button_guard
    def not_found(self):
        self.cancel_auto_timer()
        self.close_pause_dialog()
        if self.auto_run and self.auto_run_paused:
            self.auto_run_paused = False

        if not self.current_pdf:
            return
        project_id = self.correct_entry.get().strip()
        if not project_id:
            project_id = self.extract_project_id(self.current_text) or ''
        if project_id:
            new_name = settings.build_filename('not_found', project_id=project_id)
            messagebox.showinfo("提示", f"文件将重命名为“{new_name}”，后续自动运行时会自动尝试归档到对应项目文件夹。")
        else:
            new_name = settings.build_filename('unidentified', original=os.path.basename(self.current_pdf))
        new_path = os.path.join(os.path.dirname(self.current_pdf), new_name)

        try:
            self.perform_file_operation(self.current_pdf, new_path, is_copy=False)
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.append_record(self.not_found_excel, [timestamp, os.path.basename(self.current_pdf),
                                                       new_name, project_id])
            self.pdf_files.pop(self.current_index)
            self.listbox.delete(self.current_index)
            self.current_pdf = None
            self.canvas.delete("all")
            self.original_preview_img = None
            self.folder_combo.set('')
            self.correct_entry.delete(0, tk.END)
            self.status_var.set("已标记为未找到")
            if self.auto_run:
                self.handle_auto_run_after_action(removed=True)
        except Exception as e:
            messagebox.showerror("重命名失败", str(e))

    @button_guard
    def skip_pdf(self):
        self.cancel_auto_timer()
        self.close_pause_dialog()
        if self.auto_run and self.auto_run_paused:
            self.auto_run_paused = False

        if not self.current_pdf:
            return

        try:
            old_path = self.current_pdf
            old_name = os.path.basename(old_path)
            dir_name = os.path.dirname(old_path)

            project_id = self.extract_project_id(self.current_text)
            new_name = None

            if project_id:
                if not self.matched_folders:
                    new_name = settings.build_filename('skip_identified', project_id=project_id)
            else:
                new_name = settings.build_filename('unidentified', original=old_name)

            if new_name:
                new_path = os.path.join(dir_name, new_name)
                counter = 1
                base, ext = os.path.splitext(new_path)
                while os.path.exists(new_path):
                    new_path = os.path.join(dir_name, f"{base}_{counter}{ext}")
                    counter += 1

                self.perform_file_operation(old_path, new_path, is_copy=False)

                idx = self.pdf_files.index(old_path)
                self.pdf_files[idx] = new_path
                self.listbox.delete(idx)
                self.listbox.insert(idx, os.path.basename(new_path))
                self.listbox.selection_set(idx)

                self.current_pdf = new_path
                self.status_var.set(f"已重命名为: {os.path.basename(new_path)}")
        except Exception as e:
            messagebox.showerror("重命名失败", str(e))
            return

        if self.auto_run:
            if self.current_index < len(self.pdf_files):
                self.pdf_files.pop(self.current_index)
                self.listbox.delete(self.current_index)
            self.current_pdf = None
            self.canvas.delete("all")
            self.original_preview_img = None
            self.folder_combo.set('')
            self.correct_entry.delete(0, tk.END)
            self.status_var.set("已跳过")
            self.handle_auto_run_after_action(removed=True)
        else:
            self.current_pdf = None
            self.canvas.delete("all")
            self.original_preview_img = None
            self.folder_combo.set('')
            self.correct_entry.delete(0, tk.END)
            self.status_var.set("已跳过")
            cur = self.listbox.curselection()
            if cur:
                next_idx = cur[0] + 1
                if next_idx < self.listbox.size():
                    self.listbox.selection_clear(0, tk.END)
                    self.listbox.selection_set(next_idx)
                    self.listbox.activate(next_idx)
                    self.on_select_pdf(None)
                else:
                    self.status_var.set("已到最后一个文件")

    def on_closing(self):
        if self.auto_timer_id:
            self.root.after_cancel(self.auto_timer_id)
            self.auto_timer_id = None
        if self.search_timer:
            self.root.after_cancel(self.search_timer)
            self.search_timer = None
        self.indexing = False
        if self.index_thread and self.index_thread.is_alive():
            self.index_thread.join(timeout=2)
        self.save_excel_files()
        if self.has_new_record:
            extra_msg = ""
            if self.no_intervene_var.get():
                extra_msg = "\n\n⚠️ 您开启了“不干预模式”，请务必查看「未找到项目」Excel表！"
            response = messagebox.askyesnocancel("记录未导出",
                f"本次操作已有记录生成。\n是否打开记录文件夹查看？{extra_msg}\n\n（选择“是”打开文件夹，选择“否”直接退出，选择“取消”留在程序）")
            if response is None:
                return
            elif response:
                os.startfile(self.records_folder)
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = PDFMoverApp(root)
    root.mainloop()